from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session, joinedload
from io import BytesIO
from openpyxl import Workbook, load_workbook
from app.database import get_db, DATABASE_URL
from app.models import Client, Order, Task, WorkReport, Attendance, User, Product, Service
from app.auth import get_current_user, require_admin
from app.audit import log
import logging

logger = logging.getLogger("crm.import")

router = APIRouter(prefix="/api/export", tags=["export"], dependencies=[Depends(get_current_user)])


def _stream(wb: Workbook, filename: str):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/clients")
def export_clients(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["ID", "ФИО", "Телефон", "Дата создания"])
    for c in db.query(Client).all():
        ws.append([c.id, c.full_name, c.phone, c.created_at.isoformat()])
    return _stream(wb, "clients.xlsx")


@router.get("/orders")
def export_orders(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    ws.append(["ID", "Клиент", "Тип", "Статус", "Сумма", "Дата"])
    orders = db.query(Order).options(joinedload(Order.client)).order_by(Order.created_at.desc()).all()
    for o in orders:
        ws.append([o.id, o.client_name or o.client.full_name, o.order_type, o.status, o.total_price, o.created_at.isoformat()])
    return _stream(wb, "orders.xlsx")


@router.get("/reports")
def export_reports(db: Session = Depends(get_db), _=Depends(require_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёты"
    ws.append(["ID", "Сотрудник", "Работа", "Кол-во", "Цена", "Сумма", "Дата"])
    reports = db.query(WorkReport).options(
        joinedload(WorkReport.user), joinedload(WorkReport.task)
    ).order_by(WorkReport.created_at.desc()).all()
    for r in reports:
        ws.append([r.id, r.user.full_name or r.user.username, r.task.name, r.quantity, r.task.price, r.quantity * r.task.price, r.created_at.isoformat()])
    return _stream(wb, "reports.xlsx")


@router.get("/attendance")
def export_attendance(db: Session = Depends(get_db), _=Depends(require_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отметки"
    ws.append(["ID", "Сотрудник", "Дата", "Пришёл", "Ушёл"])
    records = db.query(Attendance).options(joinedload(Attendance.user)).order_by(Attendance.created_at.desc()).all()
    for r in records:
        ws.append([r.id, r.user.full_name or r.user.username, r.date, r.check_in, r.check_out or ""])
    return _stream(wb, "attendance.xlsx")


@router.get("/warehouse")
def export_warehouse(db: Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Склад"
    ws.append(["ID", "Название", "Цвет", "Артикул", "Количество"])
    for p in db.query(Product).order_by(Product.name).all():
        ws.append([p.id, p.name, p.color, p.article, p.quantity])
    return _stream(wb, "warehouse.xlsx")


@router.get("/database")
def export_database(_=Depends(require_admin)):
    db_path = DATABASE_URL.replace("sqlite:///", "")
    return FileResponse(
        db_path,
        media_type="application/octet-stream",
        filename="crm_backup.db",
        headers={"Content-Disposition": "attachment; filename=crm_backup.db"},
    )


@router.post("/import/clients")
def import_clients(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    wb = load_workbook(file.file)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            name, phone = str(row[0]).strip(), str(row[1]).strip()
            existing = db.query(Client).filter(Client.full_name == name, Client.phone == phone).first()
            if not existing:
                db.add(Client(full_name=name, phone=phone))
                count += 1
    db.commit()
    log(user, "import", "client", None, f"Импортировано клиентов: {count}", db=db)
    return {"imported": count}


@router.post("/import/services")
def import_services(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    wb = load_workbook(file.file)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            name, price, category = str(row[0]).strip(), float(row[1]), str(row[2]).strip() if row[2] else "repair"
            existing = db.query(Service).filter(Service.name == name).first()
            if not existing:
                db.add(Service(name=name, price=price, category=category))
                count += 1
    db.commit()
    log(user, "import", "service", None, f"Импортировано услуг: {count}", db=db)
    return {"imported": count}


@router.post("/import/warehouse")
def import_warehouse(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    wb = load_workbook(file.file)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            name = str(row[0]).strip()
            color = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            article = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            qty = int(row[3]) if len(row) > 3 and row[3] else 0
            existing = db.query(Product).filter(Product.article == article).first() if article else None
            if not existing:
                db.add(Product(name=name, color=color, article=article, quantity=qty))
                count += 1
            elif article:
                existing.quantity += qty
    db.commit()
    log(user, "import", "product", None, f"Импортировано товаров: {count}", db=db)
    return {"imported": count}
