import hashlib
import io
import math
import os
import secrets
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timedelta
import json
from contextlib import asynccontextmanager
from pathlib import Path

from fpdf import FPDF

import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, Request, Form, HTTPException, Depends, Query, Body, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from jose import jwt
from sqlalchemy import (
    create_engine, String, Float, Text, DateTime, Integer, Boolean, ForeignKey,
    func, select, desc, event, or_, text,
)
from sqlalchemy.orm import (
    DeclarativeBase, Mapped, mapped_column, relationship,
    Session, joinedload,
)

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
DB_URL = os.getenv("DATABASE_URL", f"sqlite:///{BASE_DIR / 'repair_crm.db'}")
if DB_URL.startswith("sqlite:///"):
    DB_URL = f"sqlite:///{BASE_DIR / DB_URL.replace('sqlite:///', '')}"
SECRET_KEY = os.getenv("SECRET_KEY", "default-secret-change-me")
TOKEN_EXPIRY = 30 * 24 * 3600

engine = create_engine(DB_URL, connect_args={"check_same_thread": False})


@event.listens_for(engine, "connect")
def _set_wal(dbapi_connection, _connection_record):
    dbapi_connection.execute("PRAGMA journal_mode=WAL")


def get_db():
    with Session(engine) as s:
        yield s


class Base(DeclarativeBase):
    pass


# ══════════════════════════════════════════════════════════════════
#  Models
# ══════════════════════════════════════════════════════════════════

class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(primary_key=True)
    username: Mapped[str] = mapped_column(String(100), unique=True)
    password_hash: Mapped[str] = mapped_column(String(200))
    full_name: Mapped[str] = mapped_column(String(200))
    role_id: Mapped[int] = mapped_column(ForeignKey("roles.id"), index=True)
    role = relationship("Role")
    inn: Mapped[str] = mapped_column(String(20), default="")
    position: Mapped[str] = mapped_column(String(100), default="")
    is_active: Mapped[bool] = mapped_column(Boolean, default=True)
    last_login: Mapped[datetime | None] = mapped_column(DateTime, default=None, nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Role(Base):
    __tablename__ = "roles"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(100), unique=True)
    permissions: Mapped[str] = mapped_column(Text, default="[]")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class AuditLog(Base):
    __tablename__ = "audit_logs"
    id: Mapped[int] = mapped_column(primary_key=True)
    user_id: Mapped[int | None] = mapped_column(ForeignKey("users.id"), nullable=True, index=True)
    user_name: Mapped[str] = mapped_column(String(200), default="")
    action: Mapped[str] = mapped_column(String(100))
    entity_type: Mapped[str] = mapped_column(String(50))
    entity_id: Mapped[int | None] = mapped_column(nullable=True)
    details: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Client(Base):
    __tablename__ = "clients"
    id: Mapped[int] = mapped_column(primary_key=True)
    full_name: Mapped[str] = mapped_column(String(200))
    phone: Mapped[str] = mapped_column(String(50))
    comment: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    orders = relationship("Order", back_populates="client", cascade="all, delete-orphan")


class Service(Base):
    __tablename__ = "services"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(200))
    price: Mapped[float] = mapped_column(Float, default=0)
    description: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Part(Base):
    __tablename__ = "parts"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(200))
    article: Mapped[str] = mapped_column(String(100), unique=True)
    purchase_price: Mapped[float] = mapped_column(Float, default=0)
    quantity: Mapped[int] = mapped_column(Integer, default=0)
    min_stock: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    movements = relationship("StockMovement", back_populates="part", cascade="all, delete-orphan",
                             order_by="StockMovement.id")


class StockMovement(Base):
    __tablename__ = "stock_movements"
    id: Mapped[int] = mapped_column(primary_key=True)
    part_id: Mapped[int] = mapped_column(ForeignKey("parts.id"), index=True)
    part = relationship("Part", back_populates="movements")
    type: Mapped[str] = mapped_column(String(10))
    quantity: Mapped[int] = mapped_column(Integer, default=0)
    price_per_unit: Mapped[float] = mapped_column(Float, default=0)
    reason: Mapped[str] = mapped_column(String(500), default="")
    order_id: Mapped[int | None] = mapped_column(ForeignKey("orders.id"), nullable=True, index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Product(Base):
    __tablename__ = "products"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(200))
    article: Mapped[str] = mapped_column(String(100), unique=True)
    color: Mapped[str] = mapped_column(String(100), default="")
    quantity: Mapped[int] = mapped_column(Integer, default=0)
    cost_price: Mapped[float] = mapped_column(Float, default=0)
    image: Mapped[str] = mapped_column(String(300), default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    movements = relationship("ProductMovement", back_populates="product",
                             cascade="all, delete-orphan", order_by="ProductMovement.id")


class ProductMovement(Base):
    __tablename__ = "product_movements"
    id: Mapped[int] = mapped_column(primary_key=True)
    product_id: Mapped[int] = mapped_column(ForeignKey("products.id"), index=True)
    product = relationship("Product", back_populates="movements")
    type: Mapped[str] = mapped_column(String(10))
    quantity: Mapped[int] = mapped_column(Integer, default=0)
    destination: Mapped[str] = mapped_column(String(50), default="")
    reason: Mapped[str] = mapped_column(String(500), default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Filament(Base):
    __tablename__ = "filaments"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(200))
    article: Mapped[str] = mapped_column(String(100), default="")
    manufacturer: Mapped[str] = mapped_column(String(100), default="")
    type: Mapped[str] = mapped_column(String(50), default="PLA")
    color: Mapped[str] = mapped_column(String(100), default="")
    quantity: Mapped[int] = mapped_column(Integer, default=0)
    grams_per_spool: Mapped[int] = mapped_column(Integer, default=1000)
    min_stock: Mapped[int] = mapped_column(Integer, default=500)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    movements = relationship("FilamentMovement", back_populates="filament",
                             cascade="all, delete-orphan", order_by="FilamentMovement.id")


class FilamentMovement(Base):
    __tablename__ = "filament_movements"
    id: Mapped[int] = mapped_column(primary_key=True)
    filament_id: Mapped[int] = mapped_column(ForeignKey("filaments.id"), index=True)
    filament = relationship("Filament", back_populates="movements")
    type: Mapped[str] = mapped_column(String(10))
    quantity: Mapped[int] = mapped_column(Integer, default=0)
    reason: Mapped[str] = mapped_column(String(500), default="")
    order_id: Mapped[int | None] = mapped_column(ForeignKey("orders.id"), nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class PrintJob(Base):
    __tablename__ = "print_jobs"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(300))
    filament_id: Mapped[int] = mapped_column(ForeignKey("filaments.id"), index=True)
    filament = relationship("Filament")
    created_by: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    creator = relationship("User")
    grams: Mapped[int] = mapped_column(Integer, default=0)
    hours: Mapped[float] = mapped_column(Float, default=0)
    status: Mapped[str] = mapped_column(String(10), default="success")
    waste_grams: Mapped[int] = mapped_column(Integer, default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class ChatMessage(Base):
    __tablename__ = "chat_messages"
    id: Mapped[int] = mapped_column(primary_key=True)
    from_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    from_user = relationship("User", foreign_keys=[from_user_id])
    to_user_id: Mapped[int | None] = mapped_column(ForeignKey("users.id"), nullable=True, index=True)
    to_user = relationship("User", foreign_keys=[to_user_id])
    text: Mapped[str] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Order(Base):
    __tablename__ = "orders"
    id: Mapped[int] = mapped_column(primary_key=True)
    client_id: Mapped[int] = mapped_column(ForeignKey("clients.id"), index=True)
    client = relationship("Client", back_populates="orders")
    order_type: Mapped[str] = mapped_column(String(20), default="repair")
    printer: Mapped[str] = mapped_column(String(200))
    defect: Mapped[str] = mapped_column(Text)
    status: Mapped[str] = mapped_column(String(20), default="in_progress")
    total_price: Mapped[float] = mapped_column(Float, default=0)
    assigned_to: Mapped[int | None] = mapped_column(ForeignKey("users.id"), nullable=True, index=True)
    assignee = relationship("User")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    closed_at: Mapped[datetime | None] = mapped_column(DateTime, default=None, nullable=True)
    deadline: Mapped[datetime | None] = mapped_column(DateTime, default=None, nullable=True)
    scheduled_at: Mapped[datetime | None] = mapped_column(DateTime, default=None, nullable=True)
    schedule_location: Mapped[str] = mapped_column(String(300), default="")
    is_confirmed: Mapped[bool] = mapped_column(Boolean, default=False)
    warranty_days: Mapped[int] = mapped_column(Integer, default=0)
    is_warranty: Mapped[bool] = mapped_column(Boolean, default=False)
    prepaid: Mapped[float] = mapped_column(Float, default=0)
    items = relationship(
        "OrderItem", back_populates="order",
        cascade="all, delete-orphan", order_by="OrderItem.id",
    )
    parts = relationship(
        "OrderPart", back_populates="order",
        cascade="all, delete-orphan", order_by="OrderPart.id",
    )


class OrderItem(Base):
    __tablename__ = "order_items"
    id: Mapped[int] = mapped_column(primary_key=True)
    order_id: Mapped[int] = mapped_column(ForeignKey("orders.id"), index=True)
    order = relationship("Order", back_populates="items")
    name: Mapped[str] = mapped_column(String(300))
    price: Mapped[float] = mapped_column(Float, default=0)


class OrderPart(Base):
    __tablename__ = "order_parts"
    id: Mapped[int] = mapped_column(primary_key=True)
    order_id: Mapped[int] = mapped_column(ForeignKey("orders.id"), index=True)
    order = relationship("Order", back_populates="parts")
    part_id: Mapped[int] = mapped_column(ForeignKey("parts.id"), index=True)
    part = relationship("Part")
    quantity: Mapped[int] = mapped_column(Integer, default=1)
    price: Mapped[float] = mapped_column(Float, default=0)


class Task(Base):
    __tablename__ = "tasks"
    id: Mapped[int] = mapped_column(primary_key=True)
    title: Mapped[str] = mapped_column(String(300))
    description: Mapped[str] = mapped_column(Text, default="")
    created_by: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    creator = relationship("User", foreign_keys=[created_by])
    assigned_to: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    assignee = relationship("User", foreign_keys=[assigned_to])
    status: Mapped[str] = mapped_column(String(20), default="pending")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    completed_at: Mapped[datetime | None] = mapped_column(DateTime, default=None, nullable=True)


# ══════════════════════════════════════════════════════════════════
#  App Setup
# ══════════════════════════════════════════════════════════════════

class Attendance(Base):
    __tablename__ = "attendance"
    id: Mapped[int] = mapped_column(primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    user = relationship("User")
    date_str: Mapped[str] = mapped_column(String(10), default="")
    check_in: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())
    check_out: Mapped[datetime | None] = mapped_column(DateTime, default=None, nullable=True)
    report: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Schedule(Base):
    __tablename__ = "schedules"
    id: Mapped[int] = mapped_column(primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    user = relationship("User")
    date: Mapped[datetime] = mapped_column(DateTime)
    time_from: Mapped[str] = mapped_column(String(5))
    time_to: Mapped[str] = mapped_column(String(5))
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


class Notification(Base):
    __tablename__ = "notifications"
    id: Mapped[int] = mapped_column(primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    user = relationship("User")
    title: Mapped[str] = mapped_column(String(200))
    text: Mapped[str] = mapped_column(Text, default="")
    link: Mapped[str] = mapped_column(String(300), default="")
    is_read: Mapped[bool] = mapped_column(Boolean, default=False)
    created_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now())


@asynccontextmanager
async def lifespan(app: FastAPI):
    Base.metadata.create_all(bind=engine)
    # DB migrations
    with engine.connect() as conn:
        for col, dtype in [("order_type", "VARCHAR(20) DEFAULT 'repair'"),
                           ("scheduled_at", "DATETIME"), ("schedule_location", "VARCHAR(300) DEFAULT ''"),
                           ("is_confirmed", "BOOLEAN DEFAULT 0"),
                           ("warranty_days", "INTEGER DEFAULT 0"),
                            ("is_warranty", "BOOLEAN DEFAULT 0"),
                            ("prepaid", "FLOAT DEFAULT 0")]:
            try:
                conn.execute(text(f"ALTER TABLE orders ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("report", "TEXT DEFAULT ''")]:
            try:
                conn.execute(text(f"ALTER TABLE attendance ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("last_login", "DATETIME")]:
            try:
                conn.execute(text(f"ALTER TABLE users ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("to_user_id", "INTEGER")]:
            try:
                conn.execute(text(f"ALTER TABLE chat_messages ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("status", "VARCHAR(10) DEFAULT 'success'"),
                            ("waste_grams", "INTEGER DEFAULT 0")]:
            try:
                conn.execute(text(f"ALTER TABLE print_jobs ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("cost_price", "FLOAT DEFAULT 0")]:
            try:
                conn.execute(text(f"ALTER TABLE products ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("date_str", "VARCHAR(10) DEFAULT ''")]:
            try:
                conn.execute(text(f"ALTER TABLE attendance ADD COLUMN {col} {dtype}"))
                conn.commit()
                conn.execute(text(
                    "UPDATE attendance SET date_str = strftime('%Y-%m-%d', date) WHERE date_str = ''"
                ))
                conn.commit()
            except Exception:
                pass
        for col, dtype in [("article", "VARCHAR(100) DEFAULT ''"),
                            ("manufacturer", "VARCHAR(100) DEFAULT ''"),
                            ("grams_per_spool", "INTEGER DEFAULT 1000")]:
            try:
                conn.execute(text(f"ALTER TABLE filaments ADD COLUMN {col} {dtype}"))
                conn.commit()
            except Exception:
                pass
        # Create notifications table if not exists
        with engine.connect() as nc:
            nc.execute(text("""
                CREATE TABLE IF NOT EXISTS notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL REFERENCES users(id),
                    title VARCHAR(200) NOT NULL,
                    text TEXT DEFAULT '',
                    link VARCHAR(300) DEFAULT '',
                    is_read BOOLEAN DEFAULT 0,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            nc.commit()
        # Create filaments table if not exists
        with engine.connect() as fc:
            fc.execute(text("""
                CREATE TABLE IF NOT EXISTS filaments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(200) NOT NULL,
                    type VARCHAR(50) DEFAULT 'PLA',
                    color VARCHAR(100) DEFAULT '',
                    quantity INTEGER DEFAULT 0,
                    min_stock INTEGER DEFAULT 500,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            fc.execute(text("""
                CREATE TABLE IF NOT EXISTS filament_movements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filament_id INTEGER NOT NULL REFERENCES filaments(id),
                    type VARCHAR(10) NOT NULL,
                    quantity INTEGER DEFAULT 0,
                    reason VARCHAR(500) DEFAULT '',
                    order_id INTEGER REFERENCES orders(id),
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            fc.execute(text("""
                CREATE TABLE IF NOT EXISTS print_jobs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name VARCHAR(300) NOT NULL,
                    filament_id INTEGER NOT NULL REFERENCES filaments(id),
                    created_by INTEGER NOT NULL REFERENCES users(id),
                    grams INTEGER DEFAULT 0,
                    hours FLOAT DEFAULT 0,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            fc.commit()
        # Create indexes on FK columns (silently skip if exist)
        for tbl, col in [("tasks", "assigned_to"), ("tasks", "created_by"),
                         ("attendance", "user_id"), ("order_items", "order_id"),
                         ("order_parts", "order_id"), ("orders", "client_id"),
                         ("schedules", "user_id"), ("chat_messages", "from_user_id"),
                         ("orders", "assigned_to")]:
            try:
                conn.execute(text(f"CREATE INDEX IF NOT EXISTS ix_{tbl}_{col} ON {tbl} ({col})"))
                conn.commit()
            except Exception:
                pass
    _seed_data()
    yield


app = FastAPI(title="CRM — Ремонт 3D принтеров", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
_tz_offset = int(os.getenv("TZ_OFFSET", "0"))
if _tz_offset == 0:
    try:
        import time as _time
        _tz_offset = 3 + (_time.timezone // 3600)
    except Exception:
        _tz_offset = 3
TIMEZONE_OFFSET = timedelta(hours=_tz_offset)
templates.env.filters["money"] = lambda x: f"{x:,.0f}".replace(",", " ") + " ₽"
templates.env.filters["dt"] = lambda x: (x + TIMEZONE_OFFSET).strftime("%d.%m.%Y %H:%M") if x else "—"
templates.env.filters["tm"] = lambda x: (x + TIMEZONE_OFFSET).strftime("%H:%M") if x else "—"
templates.env.filters["int"] = lambda x: f"{x:,}".replace(",", " ") if x else "0"
_MONTHS_RU = ["","январь","февраль","март","апрель","май","июнь","июль","август","сентябрь","октябрь","ноябрь","декабрь"]
templates.env.filters["month_ru"] = lambda dt: _MONTHS_RU[(dt + TIMEZONE_OFFSET).month] if dt else "—"

UPLOADS_DIR = BASE_DIR / "static" / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


# ══════════════════════════════════════════════════════════════════
#  Auth Helpers
# ══════════════════════════════════════════════════════════════════

def _hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100_000)
    return salt + ":" + dk.hex()


def _verify_password(password: str, stored: str) -> bool:
    salt, dk_hex = stored.split(":", 1)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100_000)
    return dk.hex() == dk_hex


def _create_token(user_id: int) -> str:
    return jwt.encode(
        {"sub": str(user_id), "exp": datetime.utcnow().timestamp() + TOKEN_EXPIRY},
        SECRET_KEY, algorithm="HS256",
    )


def _get_user_from_request(request: Request) -> User | None:
    token = request.cookies.get("token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        user_id = int(payload["sub"])
    except Exception:
        return None
    with Session(engine) as s:
        return s.execute(
            select(User).options(joinedload(User.role)).where(User.id == user_id)
        ).unique().scalar_one_or_none()


def _has_permission(user: User | None, perm: str) -> bool:
    if not user or not user.is_active:
        return False
    if user.role.name == "admin":
        return True
    perms = json.loads(user.role.permissions) if user.role.permissions else []
    return perm in perms


def _check_perm(request: Request, perm: str):
    if not _has_permission(request.state.user, perm):
        raise HTTPException(403, "Недостаточно прав")


def _validate_password(pw: str) -> str:
    pw = pw.strip()
    if " " in pw:
        raise HTTPException(400, "Пароль не должен содержать пробелы")
    if len(pw) < 4:
        raise HTTPException(400, "Пароль должен быть не короче 4 символов")
    return pw


def _audit(action: str, entity_type: str, entity_id: int | None = None,
           details: str = "", user: User | None = None, session: Session | None = None):
    a = AuditLog(
        user_id=user.id if user else None,
        user_name=user.full_name if user else "Система",
        action=action, entity_type=entity_type,
        entity_id=entity_id, details=details,
    )
    if session is not None:
        session.add(a)
        try:
            session.commit()
        except Exception:
            pass
    else:
        with Session(engine) as s:
            s.add(a)
            s.commit()
    _notify_all(f"{action}: {entity_type}", details or "", "", session)


def _notify(user_id: int, title: str, text: str = "", link: str = "", session: Session | None = None):
    n = Notification(user_id=user_id, title=title, text=text, link=link)
    if session is not None:
        session.add(n)
    else:
        with Session(engine) as s:
            s.add(n)
            s.commit()


def _notify_all(title: str, text: str = "", link: str = "", session: Session | None = None):
    with Session(engine) as s2:
        users = s2.execute(select(User.id).where(User.is_active == True)).scalars().all()
    for uid in users:
        _notify(uid, title, text, link, session)


def _seed_data():
    with Session(engine) as s:
        if s.execute(select(func.count(User.id))).scalar() > 0:
            return
        admin_role = Role(name="admin", permissions='["*"]')
        manager_role = Role(name="manager", permissions=json.dumps([
            "manage_clients", "manage_services", "manage_orders",
            "manage_warehouse", "manage_products",
        ]))
        worker_role = Role(name="worker", permissions=json.dumps([
            "manage_orders", "manage_warehouse", "manage_products",
        ]))
        s.add_all([admin_role, manager_role, worker_role])
        s.flush()
        s.add(User(
            username="admin", password_hash=_hash_password("admin"),
            full_name="Администратор", role_id=admin_role.id,
            inn="770000000000", position="Старший мастер",
        ))
        # Test clients
        clients = [
            Client(full_name="Иван Петров", phone="+7 (999) 123-45-67", comment="Постоянный клиент"),
            Client(full_name="Сергей Иванов", phone="+7 (916) 555-33-22", comment=""),
            Client(full_name="Анна Смирнова", phone="+7 (903) 777-88-99", comment="Студия 3D-печати"),
            Client(full_name="ООО «Прототип»", phone="+7 (495) 111-22-33", comment="Юр. лицо, договор №12"),
            Client(full_name="Дмитрий Козлов", phone="+7 (926) 444-55-66", comment="Срочные ремонты"),
        ]
        s.add_all(clients)
        # Test services
        services = [
            Service(name="Диагностика", price=500, description="Полная проверка принтера"),
            Service(name="Замена хотэнда", price=1500, description="Замена нагревательного блока"),
            Service(name="Чистка сопла", price=300, description="Механическая чистка засора"),
            Service(name="Замена терморезистора", price=800, description="NTC 100K"),
            Service(name="Ремонт платы управления", price=3500, description="Диагностика и пайка"),
            Service(name="Калибровка стола", price=600, description="Ручная + авто"),
            Service(name="Замена ремня осей", price=1200, description="GT2 6мм"),
            Service(name="Прошивка Marlin", price=1000, description="Обновление прошивки"),
        ]
        s.add_all(services)
        # Test print services
        print_services = [
            Service(name="3D печать (PLA)", price=5, description="За грамм, пластик PLA"),
            Service(name="3D печать (PETG)", price=7, description="За грамм, пластик PETG"),
            Service(name="3D печать (ABS)", price=8, description="За грамм, пластик ABS"),
            Service(name="Постобработка", price=300, description="Удаление поддержек, шлифовка"),
            Service(name="3D моделирование", price=1500, description="Создание модели с нуля"),
        ]
        s.add_all(print_services)
        # Test parts
        parts = [
            Part(name="Хотэнд Ender 3", article="HT-E3-V2", purchase_price=450, quantity=15, min_stock=3),
            Part(name="Термистор NTC 100K", article="NTC-100K", purchase_price=120, quantity=30, min_stock=5),
            Part(name="Сопло 0.4мм", article="NOZ-04-BR", purchase_price=80, quantity=50, min_stock=10),
            Part(name="Ремень GT2 6мм", article="GT2-6MM", purchase_price=250, quantity=8, min_stock=2),
            Part(name="Вентилятор 40x40x10", article="FAN-4010", purchase_price=180, quantity=20, min_stock=5),
            Part(name="Нагревательный картридж", article="HTR-24V40", purchase_price=350, quantity=12, min_stock=3),
        ]
        s.add_all(parts)
        # Test products
        products = [
            Product(name="PLA пластик красный", article="PLA-RED-1KG", color="Красный", quantity=25),
            Product(name="PLA пластик чёрный", article="PLA-BLK-1KG", color="Чёрный", quantity=40),
            Product(name="PETG прозрачный", article="PETG-CLR-1KG", color="Прозрачный", quantity=15),
            Product(name="ABS белый", article="ABS-WHT-1KG", color="Белый", quantity=10),
        ]
        s.add_all(products)
        # Test filaments
        filaments = [
            Filament(name="PLA красный", type="PLA", color="Красный", quantity=2000, min_stock=500),
            Filament(name="PLA чёрный", type="PLA", color="Чёрный", quantity=3000, min_stock=500),
            Filament(name="PETG прозрачный", type="PETG", color="Прозрачный", quantity=1500, min_stock=300),
            Filament(name="ABS белый", type="ABS", color="Белый", quantity=1000, min_stock=300),
        ]
        s.add_all(filaments)
        s.flush()
        for f in filaments:
            s.add(FilamentMovement(filament_id=f.id, type="in", quantity=f.quantity, reason="Начальный остаток"))
        s.flush()
        # Stock in for parts
        for p in parts:
            s.add(StockMovement(part_id=p.id, type="in", quantity=p.quantity,
                                price_per_unit=p.purchase_price, reason="Начальный остаток"))
        for p in products:
            s.add(ProductMovement(product_id=p.id, type="in", quantity=p.quantity,
                                  destination="", reason="Начальный остаток"))
        s.commit()


# ══════════════════════════════════════════════════════════════════
#  Auth Middleware
# ══════════════════════════════════════════════════════════════════

PUBLIC_PATHS = {"/login", "/api/sse/events", "/api/dashboard",
                "/api/task-assignments", "/api/warehouse/products"}


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in PUBLIC_PATHS or path.startswith("/api/") or path.startswith("/static/") or path == "/favicon.ico":
        return await call_next(request)

    user = _get_user_from_request(request)
    request.state.user = user

    if not user and path != "/login":
        return RedirectResponse("/login", status_code=303)

    return await call_next(request)


# ══════════════════════════════════════════════════════════════════
#  Exception Handlers
# ══════════════════════════════════════════════════════════════════

@app.exception_handler(404)
async def not_found(request: Request, _):
    return templates.TemplateResponse(request, "404.html", {
        "user": getattr(request.state, "user", None),
    }, status_code=404)


@app.exception_handler(400)
async def bad_request(request: Request, exc: HTTPException):
    return templates.TemplateResponse(request, "error.html", {
        "title": "Ошибка", "message": exc.detail or "Некорректный запрос",
        "user": getattr(request.state, "user", None), "unread_count": 0,
    }, status_code=400)


@app.exception_handler(403)
async def forbidden(request: Request, _):
    return templates.TemplateResponse(request, "error.html", {
        "title": "Доступ запрещён", "message": "У вас нет прав для этого действия",
        "user": getattr(request.state, "user", None), "unread_count": 0,
    }, status_code=403)


@app.exception_handler(500)
async def internal_error(request: Request, exc: Exception):
    import traceback
    traceback.print_exc()
    return templates.TemplateResponse(request, "error.html", {
        "title": "Внутренняя ошибка",
        "message": str(exc),
        "user": getattr(request.state, "user", None),
        "unread_count": 0,
    }, status_code=500)


# ══════════════════════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════════════════════

def _user_context(request: Request) -> dict:
    u = getattr(request.state, "user", None)
    if not u:
        return {"user": None, "is_admin": False, "can_manage_users": False, "unread_count": 0}
    unread = 0
    try:
        with Session(engine) as s:
            unread = s.execute(
                select(func.count(Notification.id))
                .where(Notification.user_id == u.id, Notification.is_read == False)
            ).scalar() or 0
    except Exception:
        pass
    return {
        "user": u,
        "is_admin": u.role.name == "admin",
        "can_manage_users": _has_permission(u, "manage_users"),
        "can_manage_clients": _has_permission(u, "manage_clients"),
        "can_manage_orders": _has_permission(u, "manage_orders"),
        "can_manage_warehouse": _has_permission(u, "manage_warehouse"),
        "can_manage_products": _has_permission(u, "manage_products"),
        "unread_count": unread,
    }


ORDER_STATUSES = {
    "in_progress": ("В работе", "warning text-dark"),
    "waiting_parts": ("Ожидает запчастей", "info"),
    "ready": ("Готов к выдаче", "primary"),
    "closed": ("Закрыт", "success"),
}
ORDER_TYPES = {
    "repair": ("Ремонт", "tools"),
    "print": ("3D печать", "printer"),
}
ORDER_FLOW = {
    "in_progress": ["waiting_parts", "ready", "closed"],
    "waiting_parts": ["in_progress", "ready", "closed"],
    "ready": ["in_progress", "closed"],
    "closed": ["in_progress"],
}

PER_PAGE = 20


def _paginate(session, q, page: int):
    # Count without joins/options for performance
    count_q = select(func.count()).select_from(q.subquery())
    total = session.execute(count_q).scalar() or 0
    pages = max(1, math.ceil(total / PER_PAGE))
    page = max(1, min(page, pages))
    items = session.execute(q.offset((page - 1) * PER_PAGE).limit(PER_PAGE)).unique().scalars().all()
    return items, page, pages, total


def _client_dict(c: Client) -> dict:
    return {"id": c.id, "full_name": c.full_name, "phone": c.phone, "comment": c.comment or ""}


def _recalc_total(session: Session, order_id: int):
    items_sum = session.execute(
        select(func.coalesce(func.sum(OrderItem.price), 0))
        .where(OrderItem.order_id == order_id)
    ).scalar() or 0.0
    parts_sum = session.execute(
        select(func.coalesce(func.sum(OrderPart.quantity * OrderPart.price), 0))
        .where(OrderPart.order_id == order_id)
    ).scalar() or 0.0
    order = session.get(Order, order_id)
    if order:
        order.total_price = float(items_sum) + float(parts_sum)
        session.commit()


# ══════════════════════════════════════════════════════════════════
#  Auth Routes
# ══════════════════════════════════════════════════════════════════

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", {"user": None})


@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    session: Session = Depends(get_db),
):
    user = session.execute(
        select(User).where(User.username == username.strip())
    ).scalar_one_or_none()
    if not user or not _verify_password(password, user.password_hash):
        _audit("login_failed", "user", None, f"Неудачный вход: {username.strip()}", None, session)
        return templates.TemplateResponse(request, "login.html", {
            "user": None, "error": "Неверный логин или пароль",
        }, status_code=401)
    if not user.is_active:
        return templates.TemplateResponse(request, "login.html", {
            "user": None, "error": "Учётная запись отключена",
        }, status_code=403)
    user.last_login = datetime.now()
    session.commit()
    token = _create_token(user.id)
    response = RedirectResponse("/", status_code=303)
    response.set_cookie("token", token, max_age=TOKEN_EXPIRY, httponly=True)
    _audit("login", "user", user.id, user.full_name, user, session)
    return response


@app.get("/logout")
def logout():
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie("token")
    return response


# ══════════════════════════════════════════════════════════════════
#  Pages
# ══════════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    active = session.execute(
        select(func.count(Order.id)).where(Order.status == "in_progress")
    ).scalar() or 0
    closed = session.execute(
        select(func.count(Order.id)).where(Order.status == "closed")
    ).scalar() or 0
    total_clients = session.execute(select(func.count(Client.id))).scalar() or 0
    total_services = session.execute(select(func.count(Service.id))).scalar() or 0
    total_parts = session.execute(select(func.count(Part.id))).scalar() or 0
    total_products = session.execute(select(func.count(Product.id))).scalar() or 0
    overdue = session.execute(
        select(func.count(Order.id))
        .where(Order.status != "closed", Order.deadline.isnot(None), Order.deadline < func.now())
    ).scalar() or 0
    due_soon = session.execute(
        select(func.count(Order.id))
        .where(Order.status != "closed", Order.deadline.isnot(None),
               Order.deadline >= func.now(), Order.deadline < func.now() + timedelta(days=3))
    ).scalar() or 0
    low_stock = session.execute(
        select(func.count(Part.id))
        .where(Part.quantity <= Part.min_stock, Part.min_stock > 0)
    ).scalar() or 0
    my_tasks = session.execute(
        select(func.count(Task.id))
        .where(Task.assigned_to == u.id, Task.status == "pending")
    ).scalar() or 0
    total_tasks = session.execute(
        select(func.count(Task.id)).where(Task.status == "pending")
    ).scalar() or 0
    recent = session.execute(
        select(Order).options(joinedload(Order.client))
        .order_by(desc(Order.created_at)).limit(15)
    ).unique().scalars().all()
    my_recent_tasks = session.execute(
        select(Task).options(joinedload(Task.creator), joinedload(Task.assignee))
        .where(Task.assigned_to == u.id, Task.status == "pending")
        .order_by(desc(Task.created_at)).limit(5)
    ).unique().scalars().all()
    recent_logs = session.execute(
        select(AuditLog).order_by(desc(AuditLog.created_at)).limit(10)
    ).scalars().all() if u.role.name == "admin" else []
    return templates.TemplateResponse(request, "index.html", {
        **_user_context(request),
        "active_orders": active, "closed_orders": closed,
        "total_clients": total_clients, "total_services": total_services,
        "total_parts": total_parts, "total_products": total_products,
        "low_stock": low_stock, "overdue": overdue, "due_soon": due_soon,
        "my_tasks": my_tasks, "total_tasks": total_tasks,
        "my_recent_tasks": my_recent_tasks,
        "recent_orders": recent, "recent_logs": recent_logs,
        "ORDER_STATUSES": ORDER_STATUSES, "ORDER_TYPES": ORDER_TYPES,
        "datetime": datetime,
        "last_backup": (datetime.fromtimestamp((BASE_DIR / "repair_crm.db").stat().st_mtime)
                        if (BASE_DIR / "repair_crm.db").exists() else None),
        "backup_days": ((datetime.now() - datetime.fromtimestamp((BASE_DIR / "repair_crm.db").stat().st_mtime)).days
                        if (BASE_DIR / "repair_crm.db").exists() else 0),
    })


@app.get("/clients", response_class=HTMLResponse)
def clients_page(request: Request, session: Session = Depends(get_db)):
    clients = session.execute(
        select(Client).order_by(desc(Client.created_at))
    ).scalars().all()
    return templates.TemplateResponse(request, "clients.html", {
        **_user_context(request),
        "clients": clients,
        "clients_data": [_client_dict(c) for c in clients],
    })


@app.get("/services", response_class=HTMLResponse)
def services_page(request: Request, session: Session = Depends(get_db)):
    services = session.execute(
        select(Service).order_by(Service.name)
    ).scalars().all()
    return templates.TemplateResponse(request, "services.html", {
        **_user_context(request),
        "services": services,
        "services_data": [{
            "id": s.id, "name": s.name, "price": s.price,
            "description": s.description or "",
        } for s in services],
    })


@app.get("/warehouse", response_class=HTMLResponse)
def warehouse_page(request: Request, session: Session = Depends(get_db)):
    parts = session.execute(
        select(Part).order_by(Part.name, Part.article)
    ).scalars().all()
    movements = session.execute(
        select(StockMovement).options(joinedload(StockMovement.part))
        .order_by(desc(StockMovement.created_at)).limit(40)
    ).unique().scalars().all()
    return templates.TemplateResponse(request, "warehouse.html", {
        **_user_context(request),
        "parts": parts, "movements": movements,
        "parts_data": [{
            "id": p.id, "name": p.name, "article": p.article,
            "purchase_price": p.purchase_price, "quantity": p.quantity,
            "min_stock": p.min_stock,
        } for p in parts],
    })


@app.get("/products", response_class=HTMLResponse)
def products_page(request: Request, session: Session = Depends(get_db)):
    products = session.execute(
        select(Product).order_by(Product.name, Product.article)
    ).scalars().all()
    movements = session.execute(
        select(ProductMovement).options(joinedload(ProductMovement.product))
        .order_by(desc(ProductMovement.created_at)).limit(40)
    ).unique().scalars().all()
    return templates.TemplateResponse(request, "products.html", {
        **_user_context(request),
        "products": products, "movements": movements,
        "products_data": [{
            "id": p.id, "name": p.name, "article": p.article,
            "color": p.color or "", "quantity": p.quantity,
            "cost_price": p.cost_price or 0,
            "image": p.image or "",
        } for p in products],
    })


@app.get("/orders", response_class=HTMLResponse)
def orders_page(
    request: Request,
    status: str = Query(""),
    order_type: str = Query(""),
    sort: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    client: str = Query(""),
    page: int = Query(1),
    session: Session = Depends(get_db),
):
    base_q = select(Order).options(
        joinedload(Order.client), joinedload(Order.assignee),
    )
    if status:
        base_q = base_q.where(Order.status == status)
    else:
        base_q = base_q.where(Order.status != "closed")
    if order_type:
        base_q = base_q.where(Order.order_type == order_type)
    if date_from:
        try:
            df = datetime.strptime(date_from, "%Y-%m-%d")
            base_q = base_q.where(Order.created_at >= df)
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            base_q = base_q.where(Order.created_at < dt + timedelta(days=1))
        except ValueError:
            pass
    if client.strip():
        like = f"%{client.strip()}%"
        base_q = base_q.where(Order.client.has(Client.full_name.ilike(like)))

    q = base_q.order_by(desc(Order.created_at))
    if sort == "oldest":
        q = base_q.order_by(Order.created_at)
    elif sort == "id_desc":
        q = base_q.order_by(desc(Order.id))
    elif sort == "id_asc":
        q = base_q.order_by(Order.id)
    elif sort == "price_desc":
        q = base_q.order_by(desc(Order.total_price))
    elif sort == "client":
        q = base_q.join(Order.client).order_by(Client.full_name)
    orders, page, pages, total = _paginate(session, q, page)

    counts = {}
    for s_val in ["in_progress", "waiting_parts", "ready", "closed"]:
        counts[s_val] = session.execute(
            select(func.count(Order.id)).where(Order.status == s_val)
        ).scalar() or 0

    return templates.TemplateResponse(request, "orders.html", {
        **_user_context(request),
        "orders": orders, "current_status": status, "current_type": order_type,
        "current_sort": sort,
        "counts": counts, "page": page, "pages": pages, "total": total,
        "date_from": date_from, "date_to": date_to, "client_filter": client.strip(),
        "ORDER_STATUSES": ORDER_STATUSES, "ORDER_TYPES": ORDER_TYPES, "timedelta": timedelta,
        "now": datetime.now(),
    })


@app.get("/orders/new", response_class=HTMLResponse)
def order_create_page(request: Request, session: Session = Depends(get_db)):
    clients = session.execute(
        select(Client).order_by(Client.full_name)
    ).scalars().all()
    users = session.execute(
        select(User).where(User.is_active == True).order_by(User.full_name)
    ).scalars().all()
    return templates.TemplateResponse(request, "order_create.html", {
        **_user_context(request), "clients": clients, "users": users,
        "ORDER_TYPES": ORDER_TYPES,
    })


@app.get("/orders/{order_id}", response_class=HTMLResponse)
def order_detail_page(order_id: int, request: Request, session: Session = Depends(get_db)):
    order = session.execute(
        select(Order)
        .options(joinedload(Order.client), joinedload(Order.items),
                 joinedload(Order.parts).joinedload(OrderPart.part),
                 joinedload(Order.assignee))
        .where(Order.id == order_id)
    ).unique().scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")
    services = session.execute(
        select(Service).order_by(Service.name)
    ).scalars().all()
    parts = session.execute(
        select(Part).order_by(Part.name, Part.article)
    ).scalars().all()
    return templates.TemplateResponse(request, "order_detail.html", {
        **_user_context(request),
        "order": order, "services": services, "parts": parts,
        "clients": session.execute(
            select(Client).order_by(Client.full_name)
        ).scalars().all(),
        "users": session.execute(
            select(User).where(User.is_active == True).order_by(User.full_name)
        ).scalars().all(),
        "ORDER_STATUSES": ORDER_STATUSES, "ORDER_TYPES": ORDER_TYPES,
        "ORDER_FLOW": ORDER_FLOW, "timedelta": timedelta, "now": lambda: datetime.now(),
        "services_data": [{
            "id": s.id, "name": s.name, "price": s.price,
        } for s in services],
    })


@app.get("/users", response_class=HTMLResponse)
def users_page(request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u or u.role.name != "admin":
        raise HTTPException(403)
    users = session.execute(
        select(User).options(joinedload(User.role)).order_by(User.username)
    ).unique().scalars().all()
    roles = session.execute(select(Role).order_by(Role.name)).scalars().all()
    return templates.TemplateResponse(request, "users.html", {
        **_user_context(request),
        "users": users, "roles": roles,
        "user_data": [{"id": x.id, "username": x.username, "full_name": x.full_name,
                        "role_name": x.role.name, "is_active": x.is_active,
                        "inn": x.inn or "", "position": x.position or "",
                        "last_login": (x.last_login + TIMEZONE_OFFSET).strftime("%d.%m.%Y %H:%M") if x.last_login else ""}
                       for x in users],
        "roles_data": [{"id": r.id, "name": r.name, "permissions": r.permissions} for r in roles],
    })


@app.get("/audit", response_class=HTMLResponse)
def audit_page(
    request: Request,
    action: str = Query(""),
    user_id: str = Query(""),
    page: int = Query(1),
    session: Session = Depends(get_db),
):
    u = request.state.user
    if not u or u.role.name != "admin":
        raise HTTPException(403)

    q = select(AuditLog)
    if action.strip():
        q = q.where(AuditLog.action == action.strip())
    if user_id.strip():
        q = q.where(AuditLog.user_id == int(user_id.strip()))
    q = q.order_by(desc(AuditLog.created_at))

    PER_PAGE = 50
    total = session.execute(select(func.count()).select_from(q.subquery())).scalar() or 0
    pages = max(1, math.ceil(total / PER_PAGE))
    page = max(1, min(page, pages))
    logs = session.execute(q.offset((page - 1) * PER_PAGE).limit(PER_PAGE)).scalars().all()

    actions = session.execute(
        select(AuditLog.action, func.count(AuditLog.id))
        .group_by(AuditLog.action).order_by(desc(func.count(AuditLog.id)))
    ).all()

    audit_users = session.execute(
        select(User).where(User.id.in_(
            select(AuditLog.user_id).where(AuditLog.user_id.isnot(None)).distinct()
        )).order_by(User.full_name)
    ).scalars().all()

    return templates.TemplateResponse(request, "audit.html", {
        **_user_context(request),
        "logs": logs, "page": page, "pages": pages, "total": total,
        "current_action": action.strip(), "current_user_id": user_id.strip(),
        "actions": actions, "audit_users": audit_users, "timedelta": timedelta,
    })


# ══════════════════════════════════════════════════════════════════
#  Clients API
# ══════════════════════════════════════════════════════════════════

@app.post("/clients")
def create_client(
    request: Request,
    full_name: str = Form(...),
    phone: str = Form(...),
    comment: str = Form(""),
    session: Session = Depends(get_db),
):
    c = Client(full_name=full_name.strip(), phone=phone.strip(), comment=comment.strip())
    session.add(c)
    session.commit()
    _audit("create", "client", c.id, c.full_name, request.state.user, session)
    return RedirectResponse("/clients", status_code=303)


@app.post("/clients/{client_id}/edit")
def update_client(
    client_id: int, request: Request,
    full_name: str = Form(...),
    phone: str = Form(...),
    comment: str = Form(""),
    session: Session = Depends(get_db),
):
    c = session.get(Client, client_id)
    if not c:
        raise HTTPException(404)
    c.full_name = full_name.strip()
    c.phone = phone.strip()
    c.comment = comment.strip()
    session.commit()
    _audit("update", "client", c.id, c.full_name, request.state.user, session)
    return RedirectResponse("/clients", status_code=303)


@app.post("/clients/{client_id}/delete")
def delete_client(client_id: int, request: Request, session: Session = Depends(get_db)):
    c = session.get(Client, client_id)
    if not c:
        raise HTTPException(404)
    session.delete(c)
    session.commit()
    _audit("delete", "client", client_id, c.full_name, request.state.user, session)
    return RedirectResponse("/clients", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Services API
# ══════════════════════════════════════════════════════════════════

@app.post("/services")
def create_service(
    request: Request,
    name: str = Form(...),
    price: float = Form(0),
    description: str = Form(""),
    session: Session = Depends(get_db),
):
    s = Service(name=name.strip(), price=price, description=description.strip())
    session.add(s)
    session.commit()
    _audit("create", "service", s.id, s.name, request.state.user, session)
    return RedirectResponse("/services", status_code=303)


@app.post("/services/{service_id}/edit")
def update_service(
    service_id: int, request: Request,
    name: str = Form(...),
    price: float = Form(0),
    description: str = Form(""),
    session: Session = Depends(get_db),
):
    s = session.get(Service, service_id)
    if not s:
        raise HTTPException(404)
    s.name = name.strip()
    s.price = price
    s.description = description.strip()
    session.commit()
    _audit("update", "service", s.id, s.name, request.state.user, session)
    return RedirectResponse("/services", status_code=303)


@app.post("/services/{service_id}/delete")
def delete_service(service_id: int, request: Request, session: Session = Depends(get_db)):
    s = session.get(Service, service_id)
    if not s:
        raise HTTPException(404)
    session.delete(s)
    session.commit()
    _audit("delete", "service", service_id, s.name, request.state.user, session)
    return RedirectResponse("/services", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Warehouse API (Parts)
# ══════════════════════════════════════════════════════════════════

@app.post("/warehouse/receive")
async def receive_parts(request: Request, session: Session = Depends(get_db)):
    data = await request.json()
    for row in data:
        name = row.get("name", "").strip()
        article = row.get("article", "").strip()
        price = float(row.get("purchase_price", 0))
        qty = int(row.get("quantity", 0))
        if not name or not article or qty <= 0:
            continue
        existing = session.execute(
            select(Part).where(Part.article == article)
        ).scalar_one_or_none()
        if existing:
            existing.quantity += qty
            if price > 0:
                existing.purchase_price = price
            part = existing
        else:
            part = Part(name=name, article=article, purchase_price=price, quantity=qty)
            session.add(part)
            session.flush()
        session.add(StockMovement(
            part_id=part.id, type="in", quantity=qty,
            price_per_unit=price or part.purchase_price,
            reason="Приход",
        ))
        _audit("receive", "part", part.id, f"+{qty} {part.name}", request.state.user, session)
    session.commit()
    return JSONResponse({"ok": True})


@app.post("/warehouse/{part_id}/edit")
def update_part(
    part_id: int, request: Request,
    name: str = Form(...),
    purchase_price: float = Form(0),
    min_stock: int = Form(0),
    session: Session = Depends(get_db),
):
    p = session.get(Part, part_id)
    if not p:
        raise HTTPException(404)
    p.name = name.strip()
    p.purchase_price = purchase_price
    p.min_stock = min_stock
    session.commit()
    _audit("update", "part", p.id, p.name, request.state.user, session)
    return RedirectResponse("/warehouse", status_code=303)


@app.post("/warehouse/{part_id}/delete")
def delete_part(part_id: int, request: Request, session: Session = Depends(get_db)):
    p = session.get(Part, part_id)
    if not p:
        raise HTTPException(404)
    session.delete(p)
    session.commit()
    _audit("delete", "part", part_id, p.name, request.state.user, session)
    return RedirectResponse("/warehouse", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Product Warehouse API
# ══════════════════════════════════════════════════════════════════

@app.post("/products/receive")
async def receive_products(request: Request, session: Session = Depends(get_db)):
    data = await request.json()
    for row in data:
        name = row.get("name", "").strip()
        article = row.get("article", "").strip()
        color = row.get("color", "").strip()
        qty = int(row.get("quantity", 0))
        cost_price = float(row.get("cost_price", 0) or 0)
        if not name or not article or qty <= 0:
            continue
        existing = session.execute(
            select(Product).where(Product.article == article)
        ).scalar_one_or_none()
        if existing:
            existing.quantity += qty
            if color: existing.color = color
            if cost_price: existing.cost_price = cost_price
            product = existing
        else:
            product = Product(name=name, article=article, color=color, quantity=qty, cost_price=cost_price)
            session.add(product)
            session.flush()
        session.add(ProductMovement(
            product_id=product.id, type="in", quantity=qty,
            destination="", reason="Приход",
        ))
        _audit("receive", "product", product.id, f"+{qty} {product.name}", request.state.user, session)
    session.commit()
    return JSONResponse({"ok": True})


@app.post("/products/{product_id}/stock-out")
def product_stock_out(
    product_id: int, request: Request,
    quantity: int = Form(...),
    destination: str = Form(""),
    reason: str = Form(""),
    session: Session = Depends(get_db),
):
    p = session.get(Product, product_id)
    if not p:
        raise HTTPException(404)
    if p.quantity < quantity:
        raise HTTPException(400, f"Недостаточно на складе: {p.quantity} шт.")
    p.quantity -= quantity
    session.add(ProductMovement(
        product_id=product_id, type="out", quantity=quantity,
        destination=destination.strip(), reason=reason.strip(),
    ))
    session.commit()
    _audit("stock_out", "product", p.id,
           f"-{quantity} {p.name} → {destination}", request.state.user, session)
    return RedirectResponse("/products", status_code=303)


@app.post("/products/{product_id}/edit")
def update_product(
    product_id: int, request: Request,
    name: str = Form(...),
    article: str = Form(""),
    color: str = Form(""),
    cost_price: float = Form(0),
    session: Session = Depends(get_db),
):
    p = session.get(Product, product_id)
    if not p:
        raise HTTPException(404)
    p.name = name.strip()
    p.article = article.strip()
    p.color = color.strip()
    p.cost_price = float(cost_price or 0)
    session.commit()
    _audit("update", "product", p.id, p.name, request.state.user, session)
    return RedirectResponse("/products", status_code=303)


@app.post("/products/{product_id}/delete")
def delete_product(product_id: int, request: Request, session: Session = Depends(get_db)):
    p = session.get(Product, product_id)
    if not p:
        raise HTTPException(404)
    session.delete(p)
    session.commit()
    _audit("delete", "product", product_id, p.name, request.state.user, session)
    return RedirectResponse("/products", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Users & Roles API (Admin)
# ══════════════════════════════════════════════════════════════════

AVAILABLE_PERMISSIONS = [
    ("manage_users", "Управление пользователями"),
    ("manage_clients", "Управление клиентами"),
    ("manage_services", "Управление услугами"),
    ("manage_orders", "Управление заказами"),
    ("manage_warehouse", "Управление складом запчастей"),
    ("manage_products", "Управление складом товаров"),
    ("view_audit", "Просмотр аудита"),
]


@app.post("/roles")
async def create_role(
    request: Request,
    name: str = Form(...),
    session: Session = Depends(get_db),
):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    form = await request.form()
    perms = []
    for perm_key, _ in AVAILABLE_PERMISSIONS:
        if form.get(f"perm_{perm_key}") == "1":
            perms.append(perm_key)
    role = Role(name=name.strip(), permissions=json.dumps(perms))
    session.add(role)
    session.commit()
    _audit("create", "role", role.id, role.name, request.state.user, session)
    return RedirectResponse("/users", status_code=303)


@app.post("/roles/{role_id}/edit")
async def update_role(
    role_id: int, request: Request,
    name: str = Form(...),
    session: Session = Depends(get_db),
):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    role = session.get(Role, role_id)
    if not role:
        raise HTTPException(404)
    if role.name == "admin":
        raise HTTPException(400, "Нельзя редактировать роль admin")
    form = await request.form()
    perms = []
    for perm_key, _ in AVAILABLE_PERMISSIONS:
        if form.get(f"perm_{perm_key}") == "1":
            perms.append(perm_key)
    role.name = name.strip()
    role.permissions = json.dumps(perms)
    session.commit()
    _audit("update", "role", role.id, role.name, request.state.user, session)
    return RedirectResponse("/users", status_code=303)


@app.post("/roles/{role_id}/delete")
def delete_role(role_id: int, request: Request, session: Session = Depends(get_db)):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    role = session.get(Role, role_id)
    if not role:
        raise HTTPException(404)
    if role.name == "admin":
        raise HTTPException(400, "Нельзя удалить роль admin")
    # Check no users use this role
    count = session.execute(
        select(func.count(User.id)).where(User.role_id == role_id)
    ).scalar()
    if count > 0:
        raise HTTPException(400, f"Роль используется {count} пользователями")
    session.delete(role)
    session.commit()
    _audit("delete", "role", role_id, role.name, request.state.user, session)
    return RedirectResponse("/users", status_code=303)


@app.post("/users/create")
def create_user(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(...),
    role_id: int = Form(...),
    inn: str = Form(""),
    position: str = Form(""),
    session: Session = Depends(get_db),
):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    existing = session.execute(
        select(User).where(User.username == username.strip())
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(400, "Пользователь с таким логином уже существует")
    pw = _validate_password(password)
    session.add(User(
        username=username.strip(),
        password_hash=_hash_password(pw),
        full_name=full_name.strip(),
        role_id=role_id,
        inn=inn.strip(),
        position=position.strip(),
    ))
    session.commit()
    _audit("create", "user", None, full_name.strip(), request.state.user, session)
    return RedirectResponse("/users", status_code=303)


@app.post("/users/{user_id}/edit")
def update_user(
    user_id: int, request: Request,
    full_name: str = Form(...),
    role_id: int = Form(...),
    inn: str = Form(""),
    position: str = Form(""),
    password: str = Form(""),
    session: Session = Depends(get_db),
):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    u = session.get(User, user_id)
    if not u:
        raise HTTPException(404)
    u.full_name = full_name.strip()
    u.role_id = role_id
    u.inn = inn.strip()
    u.position = position.strip()
    if password.strip():
        pw = _validate_password(password)
        u.password_hash = _hash_password(pw)
    session.commit()
    _audit("update", "user", u.id, u.full_name, request.state.user, session)
    return RedirectResponse("/users", status_code=303)


@app.post("/users/{user_id}/toggle")
def toggle_user(user_id: int, request: Request, session: Session = Depends(get_db)):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    u = session.get(User, user_id)
    if not u or u.username == "admin":
        raise HTTPException(400)
    u.is_active = not u.is_active
    session.commit()
    _audit("toggle", "user", u.id, f"{u.full_name} → {'active' if u.is_active else 'inactive'}", request.state.user, session)
    return RedirectResponse("/users", status_code=303)


@app.post("/users/{user_id}/delete")
def delete_user(user_id: int, request: Request, session: Session = Depends(get_db)):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    current = request.state.user
    u = session.get(User, user_id)
    if not u:
        raise HTTPException(404)
    if u.username == "admin":
        raise HTTPException(400, "Нельзя удалить администратора")
    if u.id == current.id:
        raise HTTPException(400, "Нельзя удалить самого себя")
    session.execute(select(Task).where(
        or_(Task.created_by == user_id, Task.assigned_to == user_id)
    ).with_for_update())
    for t in session.execute(
        select(Task).where(or_(Task.created_by == user_id, Task.assigned_to == user_id))
    ).scalars().all():
        session.delete(t)
    for m in session.execute(
        select(ChatMessage).where(ChatMessage.from_user_id == user_id)
    ).scalars().all():
        session.delete(m)
    for a in session.execute(
        select(Attendance).where(Attendance.user_id == user_id)
    ).scalars().all():
        session.delete(a)
    for s in session.execute(
        select(Schedule).where(Schedule.user_id == user_id)
    ).scalars().all():
        session.delete(s)
    for o in session.execute(
        select(Order).where(Order.assigned_to == user_id)
    ).scalars().all():
        o.assigned_to = None
    name = u.full_name
    session.delete(u)
    session.commit()
    _audit("delete", "user", user_id, name, current, session)
    return RedirectResponse("/users", status_code=303)


@app.get("/profile", response_class=HTMLResponse)
def profile_page(request: Request, session: Session = Depends(get_db)):
    current = request.state.user
    if not current:
        raise HTTPException(403)
    u = session.get(User, current.id)
    return templates.TemplateResponse(request, "profile.html", {
        **_user_context(request),
        "profile": u,
    })


@app.post("/profile/edit")
def profile_edit(
    request: Request,
    full_name: str = Form(...),
    inn: str = Form(""),
    position: str = Form(""),
    password: str = Form(""),
    session: Session = Depends(get_db),
):
    current = request.state.user
    if not current:
        raise HTTPException(403)
    u = session.get(User, current.id)
    u.full_name = full_name.strip()
    u.inn = inn.strip()
    u.position = position.strip()
    if password.strip():
        pw = _validate_password(password)
        u.password_hash = _hash_password(pw)
    session.commit()
    _audit("update_profile", "user", u.id, u.full_name, u, session)
    return RedirectResponse("/profile", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Orders API
# ══════════════════════════════════════════════════════════════════

@app.post("/orders")
def create_order(
    request: Request,
    client_id: int = Form(...),
    order_type: str = Form("repair"),
    printer: str = Form(...),
    defect: str = Form(...),
    assigned_to: int = Form(0),
    deadline: str = Form(""),
    scheduled_date: str = Form(""),
    scheduled_date_print: str = Form(""),
    scheduled_time: str = Form(""),
    scheduled_at: str = Form(""),
    schedule_location: str = Form(""),
    warranty_days: int = Form(0),
    is_warranty: bool = Form(False),
    prepaid: float = Form(0),
    session: Session = Depends(get_db),
):
    if not session.get(Client, client_id):
        raise HTTPException(400, "Клиент не найден")
    deadline_val = None
    if deadline.strip():
        try:
            deadline_val = datetime.strptime(deadline.strip(), "%Y-%m-%d")
        except ValueError:
            pass
    sched_val = None
    sched_str = scheduled_at.strip() or f"{scheduled_date.strip()}T{scheduled_time.strip()}" or f"{scheduled_date_print.strip()}T00:00"
    if sched_str and sched_str != "T" and sched_str != "T00:00":
        try:
            sched_val = datetime.strptime(sched_str.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
        except ValueError:
            pass
    order = Order(
        client_id=client_id,
        order_type=order_type,
        printer=printer.strip(),
        defect=defect.strip(),
        assigned_to=assigned_to if assigned_to > 0 else None,
        deadline=deadline_val,
        scheduled_at=sched_val,
        schedule_location=schedule_location.strip(),
        warranty_days=warranty_days,
        is_warranty=is_warranty,
        prepaid=prepaid,
    )
    session.add(order)
    session.commit()
    _audit("create", "order", order.id, f"#{order.id} {printer.strip()}", request.state.user, session)
    return RedirectResponse(f"/orders/{order.id}", status_code=303)


@app.post("/orders/{order_id}/items")
def add_order_item(
    request: Request, order_id: int,
    name: str = Form(...),
    price: float = Form(0),
    session: Session = Depends(get_db),
):
    order = session.get(Order, order_id)
    if not order or order.status == "closed":
        raise HTTPException(400, "Нельзя добавить услугу в закрытый заказ")
    item = OrderItem(order_id=order_id, name=name.strip(), price=price)
    session.add(item)
    session.commit()
    _recalc_total(session, order_id)
    _audit("add_item", "order", order_id, f"+{name.strip()} {price}₽", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/items/{item_id}/delete")
def delete_order_item(
    request: Request, order_id: int, item_id: int,
    session: Session = Depends(get_db),
):
    order = session.get(Order, order_id)
    if not order or order.status == "closed":
        raise HTTPException(400)
    item = session.get(OrderItem, item_id)
    if not item or item.order_id != order_id:
        raise HTTPException(404)
    session.delete(item)
    session.commit()
    _recalc_total(session, order_id)
    _audit("remove_item", "order", order_id, f"-{item.name}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/parts")
def add_order_part(
    request: Request, order_id: int,
    part_id: int = Form(...),
    quantity: int = Form(1),
    price: float = Form(0),
    session: Session = Depends(get_db),
):
    order = session.get(Order, order_id)
    if not order or order.status == "closed":
        raise HTTPException(400, "Нельзя добавить запчасть в закрытый заказ")
    part = session.execute(
        select(Part).where(Part.id == part_id).with_for_update()
    ).scalar_one_or_none()
    if not part:
        raise HTTPException(400, "Запчасть не найдена")
    if part.quantity < quantity:
        raise HTTPException(400, f"Недостаточно на складе: {part.quantity} шт.")
    part.quantity -= quantity
    session.add(StockMovement(
        part_id=part_id, type="out", quantity=quantity,
        price_per_unit=price or part.purchase_price,
        reason=f"Заказ #{order_id}",
        order_id=order_id,
    ))
    session.add(OrderPart(
        order_id=order_id, part_id=part_id,
        quantity=quantity, price=price or part.purchase_price,
    ))
    session.commit()
    _recalc_total(session, order_id)
    _audit("add_part", "order", order_id, f"+{part.name} x{quantity}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/parts/{op_id}/delete")
def delete_order_part(
    request: Request, order_id: int, op_id: int,
    session: Session = Depends(get_db),
):
    order = session.get(Order, order_id)
    if not order or order.status == "closed":
        raise HTTPException(400)
    op = session.get(OrderPart, op_id)
    if not op or op.order_id != order_id:
        raise HTTPException(404)
    part = session.get(Part, op.part_id)
    if part:
        part.quantity += op.quantity
        session.add(StockMovement(
            part_id=op.part_id, type="in", quantity=op.quantity,
            price_per_unit=op.price,
            reason=f"Возврат из заказа #{order_id}",
            order_id=order_id,
        ))
    session.delete(op)
    session.commit()
    _recalc_total(session, order_id)
    _audit("remove_part", "order", order_id, f"-{part.name} x{op.quantity}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/close")
def close_order(order_id: int, request: Request, session: Session = Depends(get_db)):
    order = session.get(Order, order_id)
    if not order or order.status == "closed":
        raise HTTPException(400)
    order.status = "closed"
    order.closed_at = datetime.now()
    session.commit()
    _audit("close", "order", order_id, f"#{order_id} total={order.total_price}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/status")
def change_order_status(order_id: int, request: Request,
                        new_status: str = Form(...),
                        session: Session = Depends(get_db)):
    order = session.get(Order, order_id)
    if not order or order.status == "closed" or new_status == order.status:
        raise HTTPException(400)
    if new_status not in ORDER_FLOW.get(order.status, []):
        raise HTTPException(400, f"Нельзя переключить с «{ORDER_STATUSES[order.status][0]}» на «{ORDER_STATUSES[new_status][0]}»")
    order.status = new_status
    session.commit()
    _audit("change_status", "order", order_id,
           f"#{order_id} → {ORDER_STATUSES[new_status][0]}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/reopen")
def reopen_order(order_id: int, request: Request, session: Session = Depends(get_db)):
    order = session.get(Order, order_id)
    if not order or order.status != "closed":
        raise HTTPException(400)
    order.status = "in_progress"
    order.closed_at = None
    session.commit()
    _audit("reopen", "order", order_id, f"#{order_id}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/edit")
def edit_order(
    order_id: int, request: Request,
    client_id: int = Form(...),
    order_type: str = Form("repair"),
    printer: str = Form(...),
    defect: str = Form(...),
    assigned_to: int = Form(0),
    scheduled_date: str = Form(""),
    scheduled_time: str = Form(""),
    scheduled_at: str = Form(""),
    schedule_location: str = Form(""),
    warranty_days: int = Form(0),
    is_warranty: bool = Form(False),
    prepaid: float = Form(0),
    session: Session = Depends(get_db),
):
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(404)
    prev_assigned = order.assigned_to
    order.client_id = client_id
    order.order_type = order_type
    order.printer = printer.strip()
    order.defect = defect.strip()
    order.assigned_to = assigned_to if assigned_to > 0 else None
    order.prepaid = prepaid
    sched_str = scheduled_at.strip() or f"{scheduled_date.strip()}T{scheduled_time.strip()}"
    if sched_str and sched_str != "T":
        try:
            order.scheduled_at = datetime.strptime(sched_str.replace("T", " ")[:16], "%Y-%m-%d %H:%M")
        except ValueError:
            order.scheduled_at = None
    else:
        order.scheduled_at = None
    order.schedule_location = schedule_location.strip()
    order.warranty_days = warranty_days
    order.is_warranty = is_warranty
    session.commit()
    _audit("edit", "order", order_id, f"#{order_id}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/{order_id}/confirm")
def toggle_confirm(order_id: int, request: Request, session: Session = Depends(get_db)):
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(404)
    order.is_confirmed = not order.is_confirmed
    session.commit()
    _audit("confirm" if order.is_confirmed else "unconfirm", "order", order_id,
           f"#{order_id}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


@app.post("/orders/batch-close")
def batch_close_orders(request: Request, ids: str = Form(""), session: Session = Depends(get_db)):
    for oid in [int(x) for x in ids.split(",") if x.strip().isdigit()]:
        order = session.get(Order, oid)
        if order and order.status != "closed":
            order.status = "closed"
            order.closed_at = datetime.now()
            _audit("close", "order", oid, f"#{oid} (batch)", request.state.user, session)
    session.commit()
    return RedirectResponse("/orders", status_code=303)


@app.post("/orders/batch-delete")
def batch_delete_orders(request: Request, ids: str = Form(""), session: Session = Depends(get_db)):
    for oid in [int(x) for x in ids.split(",") if x.strip().isdigit()]:
        order = session.get(Order, oid)
        if order:
            for op in order.parts:
                part = session.get(Part, op.part_id)
                if part:
                    part.quantity += op.quantity
            session.delete(order)
            _audit("delete", "order", oid, f"#{oid} (batch)", request.state.user, session)
    session.commit()
    return RedirectResponse("/orders", status_code=303)


@app.get("/clients/{client_id}/orders", response_class=HTMLResponse)
def client_history(client_id: int, request: Request, session: Session = Depends(get_db)):
    client = session.get(Client, client_id)
    if not client:
        raise HTTPException(404)
    orders = session.execute(
        select(Order).options(joinedload(Order.items), joinedload(Order.parts))
        .where(Order.client_id == client_id)
        .order_by(desc(Order.created_at))
    ).unique().scalars().all()
    return templates.TemplateResponse(request, "client_history.html", {
        **_user_context(request), "client": client, "orders": orders,
        "ORDER_STATUSES": ORDER_STATUSES, "ORDER_TYPES": ORDER_TYPES, "timedelta": timedelta,
    })


@app.post("/orders/{order_id}/delete")
def delete_order(order_id: int, request: Request, session: Session = Depends(get_db)):
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(404)
    for op in order.parts:
        part = session.get(Part, op.part_id)
        if part:
            part.quantity += op.quantity
            session.add(StockMovement(
                part_id=op.part_id, type="in", quantity=op.quantity,
                price_per_unit=op.price,
                reason=f"Возврат: удаление заказа #{order_id}",
            ))
    session.delete(order)
    session.commit()
    _audit("delete", "order", order_id, f"#{order_id}", request.state.user, session)
    return RedirectResponse("/orders", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Export
# ══════════════════════════════════════════════════════════════════

def _make_excel(headers: list[str], rows: list[list], filename: str) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.get("/export/db")
def export_db(request: Request):
    if request.state.user.role.name != "admin":
        raise HTTPException(403)
    return FileResponse(BASE_DIR / "repair_crm.db", filename="repair_crm_backup.db")


@app.get("/export/clients")
def export_clients(request: Request, session: Session = Depends(get_db)):
    clients = session.execute(select(Client).order_by(Client.full_name)).scalars().all()
    rows = [[c.id, c.full_name, c.phone, c.comment, str(c.created_at)] for c in clients]
    return _make_excel(["ID", "ФИО", "Телефон", "Комментарий", "Создан"], rows, "clients.xlsx")


@app.get("/export/orders")
def export_orders(request: Request, session: Session = Depends(get_db)):
    orders = session.execute(
        select(Order).options(joinedload(Order.client))
        .order_by(desc(Order.created_at))
    ).unique().scalars().all()
    rows = [[o.id, o.client.full_name if o.client else "—", o.order_type, o.printer, o.defect, o.status,
             o.total_price, str(o.created_at), str(o.closed_at or "")] for o in orders]
    return _make_excel(["ID", "Клиент", "Тип", "Принтер", "Дефект", "Статус", "Сумма", "Создан", "Закрыт"],
                       rows, "orders.xlsx")


@app.get("/export/services")
def export_services(request: Request, session: Session = Depends(get_db)):
    services = session.execute(select(Service).order_by(Service.name)).scalars().all()
    rows = [[s.id, s.name, s.price, s.description] for s in services]
    return _make_excel(["ID", "Название", "Цена", "Описание"], rows, "services.xlsx")


@app.get("/export/parts")
def export_parts(request: Request, session: Session = Depends(get_db)):
    parts = session.execute(select(Part).order_by(Part.name)).scalars().all()
    rows = [[p.id, p.name, p.article, p.purchase_price, p.quantity, p.min_stock] for p in parts]
    return _make_excel(["ID", "Название", "Артикул", "Цена закупки", "Кол-во", "Мин. остаток"],
                       rows, "parts.xlsx")


@app.get("/export/products")
def export_products(request: Request, session: Session = Depends(get_db)):
    products = session.execute(select(Product).order_by(Product.name)).scalars().all()
    rows = [[p.id, p.name, p.article, p.color, p.quantity] for p in products]
    return _make_excel(["ID", "Название", "Артикул", "Цвет", "Кол-во"], rows, "products.xlsx")


# ══════════════════════════════════════════════════════════════════
#  Search
# ══════════════════════════════════════════════════════════════════

@app.get("/search", response_class=HTMLResponse)
def search_page(request: Request, q: str = Query(""), session: Session = Depends(get_db)):
    results = {"clients": [], "orders": [], "parts": [], "products": [], "services": []}
    if q:
        like = f"%{q.strip()}%"
        results["clients"] = session.execute(
            select(Client).where(or_(Client.full_name.ilike(like), Client.phone.ilike(like)))
            .limit(20)
        ).scalars().all()
        results["orders"] = session.execute(
            select(Order).options(joinedload(Order.client))
            .where(or_(Order.printer.ilike(like), Order.defect.ilike(like),
                       Order.client.has(Client.full_name.ilike(like))))
            .order_by(desc(Order.created_at)).limit(20)
        ).unique().scalars().all()
        results["parts"] = session.execute(
            select(Part).where(or_(Part.name.ilike(like), Part.article.ilike(like))).limit(20)
        ).scalars().all()
        results["products"] = session.execute(
            select(Product).where(or_(Product.name.ilike(like), Product.article.ilike(like))).limit(20)
        ).scalars().all()
        results["services"] = session.execute(
            select(Service).where(Service.name.ilike(like)).limit(20)
        ).scalars().all()
    return templates.TemplateResponse(request, "search.html", {
        **_user_context(request), "q": q.strip(), "results": results,
        "total": sum(len(v) for v in results.values()),
    })


# ══════════════════════════════════════════════════════════════════
#  Receipt
# ══════════════════════════════════════════════════════════════════

def _build_receipt_pdf(order, line_items) -> bytes:
    from fpdf.enums import XPos, YPos
    font_paths = [
        BASE_DIR / "arial.ttf",
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/TTF/DejaVuSans.ttf"),
    ]
    font_file = None
    for fp in font_paths:
        if fp.exists():
            font_file = str(fp)
            break

    pdf = FPDF()
    pdf.add_page()
    if font_file:
        pdf.add_font("ArialU", "", font_file)
        pdf.add_font("ArialU", "B", font_file)
        font_name = "ArialU"
    else:
        font_name = "Helvetica"

    def money(v): return f"{v:,.0f}".replace(",", " ") + " руб."

    pdf.set_font(font_name, "B", 16)
    pdf.cell(0, 10, f"Квитанция №{order.id}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font(font_name, "", 10)
    pdf.cell(0, 6, f"от {order.created_at.strftime('%d.%m.%Y')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    pdf.set_font(font_name, "", 10)
    details = [
        ("Клиент:", order.client.full_name if order.client else "—"),
        ("Телефон:", order.client.phone if order.client else "—"),
        ("Принтер:", order.printer),
        ("Дефект:", order.defect[:80]),
    ]
    if order.assignee:
        extra = f" (ИНН {order.assignee.inn})" if order.assignee.inn else ""
        details.append(("Мастер:", f"{order.assignee.full_name}{extra}"))
    for label, value in details:
        pdf.set_font(font_name, "B", 10)
        pdf.cell(28, 6, label)
        pdf.set_font(font_name, "", 10)
        pdf.cell(0, 6, value, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    col_w = [10, 78, 14, 32, 38]
    headers = ["#", "Наименование", "Кол.", "Цена", "Сумма"]
    pdf.set_font(font_name, "B", 9)
    pdf.set_fill_color(230, 230, 230)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, fill=True, align="C" if i > 0 else "C")
    pdf.ln()

    pdf.set_font(font_name, "", 9)
    for idx, (name, qty, price, line_total) in enumerate(line_items, 1):
        pdf.cell(col_w[0], 6, str(idx), border=1, align="C")
        pdf.cell(col_w[1], 6, name[:50], border=1)
        pdf.cell(col_w[2], 6, str(qty), border=1, align="C")
        pdf.cell(col_w[3], 6, money(price), border=1, align="R")
        pdf.cell(col_w[4], 6, money(line_total), border=1, align="R")
        pdf.ln()

    pdf.set_font(font_name, "B", 10)
    pdf.cell(sum(col_w[:4]), 7, "ИТОГО:", border=1, align="R")
    pdf.cell(col_w[4], 7, money(order.total_price), border=1, align="R")
    pdf.ln(10)

    if order.closed_at:
        pdf.set_font(font_name, "", 9)
        pdf.cell(0, 6, f"Выполнен: {order.closed_at.strftime('%d.%m.%Y %H:%M')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(4)

    pdf.set_font(font_name, "", 9)
    pdf.cell(85, 6, "Мастер: _________________________")
    pdf.cell(85, 6, "Клиент: _________________________", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)
    pdf.set_font(font_name, "", 8)
    pdf.cell(85, 5, "(подпись)")
    pdf.cell(85, 5, "(подпись)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return pdf.output()


@app.get("/orders/{order_id}/receipt", response_class=HTMLResponse)
def order_receipt(order_id: int, request: Request, format: str = Query("html"),
                  session: Session = Depends(get_db)):
    order = session.execute(
        select(Order)
        .options(joinedload(Order.client), joinedload(Order.items),
                 joinedload(Order.parts).joinedload(OrderPart.part),
                 joinedload(Order.assignee))
        .where(Order.id == order_id)
    ).unique().scalar_one_or_none()
    if not order:
        raise HTTPException(404)
    line_items = []
    for item in order.items:
        line_items.append((item.name, 1, item.price, item.price))
    for op in order.parts:
        if op.part:
            line_items.append((f"{op.part.name} ({op.part.article})", op.quantity, op.price, op.quantity * op.price))

    html = templates.env.get_template("receipt.html").render(
        order=order, line_items=line_items, user=None)

    if format == "pdf":
        pdf = _build_receipt_pdf(order, line_items)
        return StreamingResponse(io.BytesIO(pdf),
                                 media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename=receipt_{order.id}.pdf"})

    resp = HTMLResponse(html)
    return resp


@app.post("/orders/{order_id}/send-receipt")
def send_receipt(order_id: int, request: Request, email_to: str = Form(...),
                 session: Session = Depends(get_db)):
    order = session.execute(
        select(Order)
        .options(joinedload(Order.client), joinedload(Order.items),
                 joinedload(Order.parts).joinedload(OrderPart.part),
                 joinedload(Order.assignee))
        .where(Order.id == order_id)
    ).unique().scalar_one_or_none()
    if not order:
        raise HTTPException(404)

    line_items = []
    for item in order.items:
        line_items.append((item.name, 1, item.price, item.price))
    for op in order.parts:
        if op.part:
            line_items.append((f"{op.part.name} ({op.part.article})", op.quantity, op.price, op.quantity * op.price))

    body = templates.env.get_template("receipt_email.html").render(
        order=order, line_items=line_items)
    subject = f"Квитанция №{order.id} — Ремонт 3D принтера"

    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user)

    if not smtp_host:
        raise HTTPException(400, "SMTP не настроен. Укажите SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS в .env")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = email_to.strip()
    msg.attach(MIMEText(body, "html", "utf-8"))

    try:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=10)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_from, [email_to.strip()], msg.as_string())
        server.quit()
    except Exception as e:
        raise HTTPException(400, f"Ошибка отправки: {str(e)}")

    _audit("send_receipt", "order", order_id, f"→ {email_to.strip()}", request.state.user, session)
    return RedirectResponse(f"/orders/{order_id}", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Charts API
# ══════════════════════════════════════════════════════════════════

@app.get("/api/charts/revenue")
def chart_revenue(session: Session = Depends(get_db)):
    from datetime import date, timedelta
    today = date.today()
    months = [(today.replace(day=1) - timedelta(days=i*30)).replace(day=1) for i in range(5, -1, -1)]
    data = []
    for m in months:
        next_m = (m.replace(day=28) + timedelta(days=4)).replace(day=1)
        total = session.execute(
            select(func.coalesce(func.sum(Order.total_price), 0))
            .where(Order.status == "closed", Order.closed_at >= m, Order.closed_at < next_m)
        ).scalar() or 0
        data.append({"month": m.strftime("%b %Y"), "total": float(total)})
    return JSONResponse(data)


@app.get("/api/charts/top-services")
def chart_top_services(session: Session = Depends(get_db)):
    items = session.execute(
        select(OrderItem.name, func.count(OrderItem.id), func.sum(OrderItem.price))
        .group_by(OrderItem.name).order_by(desc(func.sum(OrderItem.price))).limit(8)
    ).all()
    return JSONResponse([{"name": r[0], "count": r[1], "total": float(r[2] or 0)} for r in items])


# ══════════════════════════════════════════════════════════════════
#  Product image upload
# ══════════════════════════════════════════════════════════════════

@app.post("/products/{product_id}/upload-image")
async def upload_product_image(product_id: int, file: UploadFile = File(...),
                               session: Session = Depends(get_db)):
    p = session.get(Product, product_id)
    if not p:
        raise HTTPException(404)
    ext = Path(file.filename).suffix or ".jpg"
    safe_name = f"prod_{product_id}_{secrets.token_hex(4)}{ext}"
    filepath = UPLOADS_DIR / safe_name
    content = await file.read()
    filepath.write_bytes(content)
    p.image = safe_name
    session.commit()
    return RedirectResponse("/products", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Filament Management
# ══════════════════════════════════════════════════════════════════

@app.get("/filaments", response_class=HTMLResponse)
def filaments_page(request: Request, session: Session = Depends(get_db)):
    filaments = session.execute(
        select(Filament).order_by(Filament.name)
    ).scalars().all()
    movements = session.execute(
        select(FilamentMovement).options(joinedload(FilamentMovement.filament))
        .order_by(desc(FilamentMovement.created_at)).limit(30)
    ).unique().scalars().all()
    return templates.TemplateResponse(request, "filaments.html", {
        **_user_context(request),
        "filaments": filaments, "movements": movements,
    })


@app.post("/filaments")
def create_filament(
    request: Request,
    name: str = Form(...),
    article: str = Form(""),
    manufacturer: str = Form(""),
    type: str = Form("PLA"),
    color: str = Form(""),
    quantity: int = Form(0),
    grams_per_spool: int = Form(1000),
    session: Session = Depends(get_db),
):
    f = Filament(name=name.strip(), article=article.strip(), manufacturer=manufacturer.strip(),
                 type=type, color=color.strip(), quantity=quantity, grams_per_spool=grams_per_spool)
    session.add(f)
    session.flush()
    if quantity > 0:
        session.add(FilamentMovement(filament_id=f.id, type="in", quantity=quantity, reason="Начальный остаток"))
    session.commit()
    return RedirectResponse("/filaments", status_code=303)


@app.post("/filaments/{fid}/edit")
def edit_filament(
    fid: int, request: Request,
    name: str = Form(...),
    article: str = Form(""),
    manufacturer: str = Form(""),
    color: str = Form(""),
    grams_per_spool: int = Form(1000),
    min_stock: int = Form(0),
    session: Session = Depends(get_db),
):
    f = session.get(Filament, fid)
    if not f: raise HTTPException(404)
    f.name = name.strip()
    f.article = article.strip()
    f.manufacturer = manufacturer.strip()
    f.color = color.strip()
    f.grams_per_spool = grams_per_spool
    f.min_stock = min_stock
    session.commit()
    return RedirectResponse("/filaments", status_code=303)


@app.post("/filaments/{fid}/delete")
def delete_filament(fid: int, request: Request, session: Session = Depends(get_db)):
    f = session.get(Filament, fid)
    if f: session.delete(f); session.commit()
    return RedirectResponse("/filaments", status_code=303)


@app.post("/filaments/receive")
async def receive_filament(request: Request, session: Session = Depends(get_db)):
    data = await request.json()
    for row in data:
        fid = int(row.get("id", 0))
        qty = int(row.get("quantity", 0))
        if fid <= 0 or qty <= 0: continue
        f = session.get(Filament, fid)
        if not f: continue
        f.quantity += qty
        session.add(FilamentMovement(filament_id=fid, type="in", quantity=qty, reason="Приход"))
    session.commit()
    return JSONResponse({"ok": True})


@app.post("/filaments/expense")
def expense_filament(
    request: Request,
    filament_id: int = Form(...),
    quantity: int = Form(...),
    reason: str = Form(""),
    session: Session = Depends(get_db),
):
    f = session.get(Filament, filament_id)
    if not f: raise HTTPException(400)
    if f.quantity < quantity: raise HTTPException(400, f"Недостаточно: {f.quantity} г.")
    f.quantity -= quantity
    session.add(FilamentMovement(filament_id=filament_id, type="out", quantity=quantity, reason=reason.strip()))
    session.commit()
    return RedirectResponse("/filaments", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Print Jobs
# ══════════════════════════════════════════════════════════════════

@app.get("/prints", response_class=HTMLResponse)
def print_jobs_page(request: Request, session: Session = Depends(get_db)):
    jobs = session.execute(
        select(PrintJob).options(joinedload(PrintJob.filament), joinedload(PrintJob.creator))
        .order_by(desc(PrintJob.created_at)).limit(100)
    ).unique().scalars().all()
    filaments = session.execute(
        select(Filament).order_by(Filament.name)
    ).scalars().all()
    total_jobs = len(jobs)
    success_jobs = sum(1 for j in jobs if j.status == "success")
    fail_jobs = sum(1 for j in jobs if j.status == "fail")
    total_grams = sum(j.grams for j in jobs)
    success_grams = sum(j.grams for j in jobs if j.status == "success")
    fail_grams = sum(j.grams for j in jobs if j.status == "fail")
    waste_grams_total = sum(j.waste_grams for j in jobs if j.status == "fail")
    return templates.TemplateResponse(request, "prints.html", {
        **_user_context(request),
        "jobs": jobs, "filaments": filaments,
        "total_jobs": total_jobs, "success_jobs": success_jobs, "fail_jobs": fail_jobs,
        "total_grams": total_grams, "success_grams": success_grams,
        "fail_grams": fail_grams, "waste_grams_total": waste_grams_total,
    })


@app.post("/prints")
def create_print_job(
    request: Request,
    name: str = Form(...),
    filament_id: int = Form(...),
    grams: int = Form(0),
    hours: float = Form(0),
    session: Session = Depends(get_db),
):
    u = request.state.user
    if not u: raise HTTPException(403)
    f = session.get(Filament, filament_id)
    if not f: raise HTTPException(400, "Пластик не найден")
    if f.quantity < grams: raise HTTPException(400, f"Недостаточно пластика: {f.quantity} г.")
    f.quantity -= grams
    session.add(FilamentMovement(filament_id=filament_id, type="out", quantity=grams, reason=f"Печать: {name.strip()}"))
    session.add(PrintJob(name=name.strip(), filament_id=filament_id, created_by=u.id, grams=grams, hours=hours))
    session.commit()
    return RedirectResponse("/prints", status_code=303)


@app.post("/prints/{job_id}/result")
def mark_print_result(job_id: int, request: Request,
                      status: str = Form(...),
                      waste_grams: int = Form(0),
                      session: Session = Depends(get_db)):
    job = session.get(PrintJob, job_id)
    if not job: raise HTTPException(404)
    job.status = status
    job.waste_grams = waste_grams if status == "fail" else 0
    session.commit()
    _audit("mark_print", "print_job", job_id,
           f"{job.name} → {'успех' if status == 'success' else 'брак'}" +
           (f", {waste_grams} г. брак" if status == "fail" else ""),
           request.state.user, session)
    return RedirectResponse("/prints", status_code=303)


@app.post("/prints/{job_id}/delete")
def delete_print_job(job_id: int, request: Request, session: Session = Depends(get_db)):
    job = session.get(PrintJob, job_id)
    if not job: raise HTTPException(404)
    f = session.get(Filament, job.filament_id)
    if f:
        f.quantity += job.grams
        session.add(FilamentMovement(filament_id=job.filament_id, type="in", quantity=job.grams,
                     reason=f"Аннулирована печать: {job.name}"))
    session.delete(job)
    session.commit()
    return RedirectResponse("/prints", status_code=303)

# ══════════════════════════════════════════════════════════════════
#  Attendance
# ══════════════════════════════════════════════════════════════════

@app.get("/attendance", response_class=HTMLResponse)
def attendance_page(request: Request, month: str = Query(""), session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    today = datetime.now() + TIMEZONE_OFFSET
    today_str = today.strftime("%Y-%m-%d")
    if month:
        try:
            year, mon = map(int, month.split("-"))
            base = datetime(year, mon, 1)
        except (ValueError, TypeError):
            base = today.replace(day=1)
    else:
        base = today.replace(day=1)
    next_month = (base.replace(day=28) + timedelta(days=4)).replace(day=1)
    prev_month = (base - timedelta(days=1)).replace(day=1)
    base_m = base.strftime("%Y-%m")
    next_m = next_month.strftime("%Y-%m")

    users = session.execute(
        select(User).where(User.is_active == True).order_by(User.full_name)
    ).scalars().all()

    attendances = session.execute(
        select(Attendance).options(joinedload(Attendance.user))
        .where(Attendance.date_str >= base_m, Attendance.date_str < next_m)
        .order_by(desc(Attendance.created_at))
    ).unique().scalars().all()

    schedules = session.execute(
        select(Schedule).options(joinedload(Schedule.user))
        .where(Schedule.date >= base, Schedule.date < next_month)
        .order_by(Schedule.date, Schedule.time_from)
    ).unique().scalars().all()

    by_user_date = {}
    today_att = {}
    for a in attendances:
        by_user_date.setdefault(a.user_id, {})[a.date_str] = a
        if a.date_str == today_str:
            today_att[a.user_id] = a

    sched_map = {}
    for s in schedules:
        d = s.date.strftime("%Y-%m-%d") if isinstance(s.date, datetime) else str(s.date)[:10]
        sched_map.setdefault(d, []).append(s)

    work_hours = {}
    for a in attendances:
        if a.check_out:
            delta = (a.check_out - a.check_in).total_seconds()
            if delta > 0:
                uid = a.user_id
                if uid not in work_hours:
                    work_hours[uid] = {"hours": 0.0, "days": 0}
                work_hours[uid]["hours"] += delta / 3600
                work_hours[uid]["days"] += 1

    return templates.TemplateResponse(request, "attendance.html", {
        **_user_context(request),
        "users": users, "base_month": base, "next_month": next_month, "prev_month": prev_month,
        "by_user_date": by_user_date, "today_att": today_att, "today": today,
        "current_user_id": u.id, "sched_map": sched_map, "timedelta": timedelta,
        "now_utc": datetime.now(), "work_hours": work_hours,
    })


@app.post("/attendance/check-in")
def check_in(request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    today_str = (datetime.now() + TIMEZONE_OFFSET).strftime("%Y-%m-%d")
    existing = session.execute(
        select(Attendance).where(Attendance.user_id == u.id, Attendance.date_str == today_str)
    ).scalar_one_or_none()
    if existing:
        return RedirectResponse("/attendance", status_code=303)
    session.add(Attendance(user_id=u.id, date_str=today_str, check_in=datetime.now()))
    session.commit()
    _audit("check_in", "attendance", None, f"{u.full_name} пришёл", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/attendance/check-out")
def check_out(request: Request, report: str = Form(""), session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    today_str = (datetime.now() + TIMEZONE_OFFSET).strftime("%Y-%m-%d")
    a = session.execute(
        select(Attendance).where(Attendance.user_id == u.id, Attendance.date_str == today_str)
    ).scalar_one_or_none()
    if a and not a.check_out:
        a.check_out = datetime.now()
        a.report = report.strip()
        session.commit()
        _audit("check_out", "attendance", None, f"{u.full_name} ушёл", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/attendance/cancel")
def cancel_check_in(request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    today_str = (datetime.now() + TIMEZONE_OFFSET).strftime("%Y-%m-%d")
    a = session.execute(
        select(Attendance).where(Attendance.user_id == u.id, Attendance.date_str == today_str)
    ).scalar_one_or_none()
    if a and not a.check_out:
        session.delete(a)
        session.commit()
        _audit("cancel_check_in", "attendance", None, f"{u.full_name} отменил", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/attendance/edit")
def edit_attendance(request: Request, check_in: str = Form(""), check_out: str = Form(""),
                    report: str = Form(""), session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    today_str = (datetime.now() + TIMEZONE_OFFSET).strftime("%Y-%m-%d")
    today_dt = (datetime.now() + TIMEZONE_OFFSET).replace(hour=0, minute=0, second=0, microsecond=0)
    a = session.execute(
        select(Attendance).where(Attendance.user_id == u.id, Attendance.date_str == today_str)
    ).scalar_one_or_none()
    if not a: raise HTTPException(400, "Нет отметки за сегодня")
    if check_in.strip():
        try:
            h, m = map(int, check_in.split(":"))
            a.check_in = today_dt.replace(hour=h, minute=m) - TIMEZONE_OFFSET
        except ValueError: pass
    if check_out.strip():
        try:
            h, m = map(int, check_out.split(":"))
            a.check_out = today_dt.replace(hour=h, minute=m) - TIMEZONE_OFFSET
        except ValueError: pass
    a.report = report.strip()
    session.commit()
    _audit("edit_attendance", "attendance", a.id, f"{u.full_name}", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/attendance/{att_id}/admin-edit")
def admin_edit_attendance(att_id: int, request: Request, check_in: str = Form(""),
                          check_out: str = Form(""), report: str = Form(""),
                          session: Session = Depends(get_db)):
    u = request.state.user
    if not u or u.role.name != "admin": raise HTTPException(403)
    a = session.get(Attendance, att_id)
    if not a: raise HTTPException(404)
    if check_in.strip():
        try:
            h, m = map(int, check_in.split(":"))
            a.check_in = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0) - TIMEZONE_OFFSET
        except ValueError: pass
    if check_out.strip():
        try:
            h, m = map(int, check_out.split(":"))
            a.check_out = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0) - TIMEZONE_OFFSET
        except ValueError: pass
    a.report = report.strip()
    session.commit()
    _audit("admin_edit_attendance", "attendance", att_id, f"{u.full_name}", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/schedule")
def save_schedule(request: Request, user_id: int = Form(0), date: str = Form(""),
                  time_from: str = Form(""), time_to: str = Form(""),
                  session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    target = session.get(User, user_id or u.id)
    if not target: raise HTTPException(400)
    try: d = datetime.strptime(date, "%Y-%m-%d")
    except ValueError: raise HTTPException(400, "Неверная дата")
    if not time_from or not time_to: raise HTTPException(400, "Укажите время")
    session.add(Schedule(user_id=target.id, date=d, time_from=time_from, time_to=time_to))
    session.commit()
    _audit("add_schedule", "schedule", None, f"{target.full_name} {date} {time_from}-{time_to}", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/schedule/bulk")
def bulk_schedule(request: Request, user_id: int = Form(0), date_from: str = Form(""),
                  date_to: str = Form(""), time_from: str = Form(""), time_to: str = Form(""),
                  workdays_only: str = Form("0"), session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    target = session.get(User, user_id or u.id)
    if not target: raise HTTPException(400)
    if not time_from or not time_to: raise HTTPException(400, "Укажите время")
    try:
        d_from = datetime.strptime(date_from, "%Y-%m-%d")
        d_to = datetime.strptime(date_to, "%Y-%m-%d")
    except ValueError: raise HTTPException(400, "Неверная дата")
    if d_to < d_from: raise HTTPException(400, "Дата «по» раньше даты «с»")
    count = 0
    cur = d_from
    while cur <= d_to:
        if workdays_only != "1" or cur.weekday() < 5:
            session.add(Schedule(user_id=target.id, date=cur, time_from=time_from, time_to=time_to))
            count += 1
        cur += timedelta(days=1)
    session.commit()
    _audit("bulk_schedule", "schedule", None, f"{target.full_name} {date_from}–{date_to} ({count} см.)", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/schedule/{sched_id}/edit")
def edit_schedule(sched_id: int, request: Request, date: str = Form(""),
                  time_from: str = Form(""), time_to: str = Form(""),
                  session: Session = Depends(get_db)):
    u = request.state.user
    if not u: raise HTTPException(403)
    s = session.get(Schedule, sched_id)
    if not s: raise HTTPException(404)
    try: d = datetime.strptime(date, "%Y-%m-%d")
    except ValueError: raise HTTPException(400, "Неверная дата")
    if not time_from or not time_to: raise HTTPException(400, "Укажите время")
    s.date = d; s.time_from = time_from; s.time_to = time_to
    session.commit()
    _audit("edit_schedule", "schedule", s.id, f"{s.user.full_name} {date} {time_from}-{time_to}", u, session)
    return RedirectResponse("/attendance", status_code=303)


@app.post("/schedule/{sched_id}/delete")
def delete_schedule(sched_id: int, request: Request, session: Session = Depends(get_db)):
    s = session.get(Schedule, sched_id)
    if s: session.delete(s); session.commit()
    return RedirectResponse("/attendance", status_code=303)

# ══════════════════════════════════════════════════════════════════
#  Chat
# ══════════════════════════════════════════════════════════════════

@app.get("/chat", response_class=HTMLResponse)
def chat_page(request: Request, peer: str = Query(""), session: Session = Depends(get_db)):
    u = request.state.user
    if not u:
        raise HTTPException(403)
    users = session.execute(
        select(User).where(User.is_active == True).order_by(User.full_name)
    ).scalars().all()
    peer_id = 0
    peer_name = "Общий чат"
    try:
        peer_id = int(peer.strip()) if peer.strip() else 0
    except ValueError:
        peer_id = 0
    q = select(ChatMessage).options(joinedload(ChatMessage.from_user), joinedload(ChatMessage.to_user))
    if peer_id == 0:
        q = q.where(ChatMessage.to_user_id.is_(None))
    else:
        q = q.where(
            or_(
                (ChatMessage.from_user_id == u.id) & (ChatMessage.to_user_id == peer_id),
                (ChatMessage.from_user_id == peer_id) & (ChatMessage.to_user_id == u.id),
            )
        )
        peer_name = (session.get(User, peer_id) or u).full_name
    q = q.order_by(desc(ChatMessage.created_at)).limit(100)
    messages = session.execute(q).unique().scalars().all()
    return templates.TemplateResponse(request, "chat.html", {
        **_user_context(request),
        "messages": list(reversed(messages)), "users": users,
        "current_user_id": u.id, "peer_id": peer_id, "peer_name": peer_name,
    })


@app.post("/chat/send")
def chat_send(request: Request, text: str = Form(...), to_user_id: int = Form(0),
              session: Session = Depends(get_db)):
    u = request.state.user
    if not u or not text.strip():
        raise HTTPException(403)
    session.add(ChatMessage(from_user_id=u.id, to_user_id=to_user_id if to_user_id > 0 else None, text=text.strip()))
    session.commit()
    redirect_url = "/chat" if to_user_id == 0 else f"/chat?peer={to_user_id}"
    return RedirectResponse(redirect_url, status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Tasks
# ══════════════════════════════════════════════════════════════════

@app.get("/tasks", response_class=HTMLResponse)
def tasks_page(
    request: Request,
    filter: str = Query("all"),
    session: Session = Depends(get_db),
):
    u = request.state.user
    if not u:
        raise HTTPException(403)

    users = session.execute(
        select(User).where(User.is_active == True).order_by(User.full_name)
    ).scalars().all()

    q = select(Task).options(
        joinedload(Task.creator), joinedload(Task.assignee),
    )
    if filter == "my":
        q = q.where(Task.assigned_to == u.id)
    elif filter == "assigned":
        q = q.where(Task.created_by == u.id, Task.assigned_to != u.id)
    elif filter == "done":
        q = q.where(Task.status == "done")
    elif filter == "pending":
        q = q.where(Task.status == "pending")
    q = q.order_by(desc(Task.created_at))

    tasks = session.execute(q).unique().scalars().all()

    counts = {}
    for f_val, f_label, f_cond in [
        ("all", "Все", True),
        ("my", "Мои", Task.assigned_to == u.id),
        ("assigned", "Назначил", (Task.created_by == u.id) & (Task.assigned_to != u.id)),
        ("pending", "Активные", Task.status == "pending"),
        ("done", "Выполнены", Task.status == "done"),
    ]:
        cnt_q = select(func.count(Task.id))
        if f_cond is not True:
            cnt_q = cnt_q.where(f_cond)
        counts[f_val] = session.execute(cnt_q).scalar() or 0

    return templates.TemplateResponse(request, "tasks.html", {
        **_user_context(request),
        "tasks": tasks, "users": users, "filter": filter, "counts": counts,
    })


@app.post("/tasks")
def create_task(
    request: Request,
    title: str = Form(...),
    description: str = Form(""),
    assigned_to: int = Form(...),
    session: Session = Depends(get_db),
):
    u = request.state.user
    if not u:
        raise HTTPException(403)
    target = session.get(User, assigned_to)
    if not target:
        raise HTTPException(400, "Пользователь не найден")
    task = Task(
        title=title.strip(),
        description=description.strip(),
        created_by=u.id,
        assigned_to=assigned_to,
    )
    session.add(task)
    session.commit()
    _audit("create", "task", task.id, f"«{task.title}» → {target.full_name}", u, session)
    return RedirectResponse("/tasks", status_code=303)


@app.post("/tasks/{task_id}/done")
def mark_task_done(task_id: int, request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u:
        raise HTTPException(403)
    task = session.get(Task, task_id)
    if not task:
        raise HTTPException(404)
    task.status = "done"
    task.completed_at = datetime.now()
    session.commit()
    _audit("done", "task", task.id, f"«{task.title}» выполнена", u, session)
    return RedirectResponse("/tasks", status_code=303)


@app.post("/tasks/{task_id}/delete")
def delete_task(task_id: int, request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u:
        raise HTTPException(403)
    task = session.get(Task, task_id)
    if not task:
        raise HTTPException(404)
    if task.created_by != u.id and u.role.name != "admin":
        raise HTTPException(403, "Удалить может только автор или администратор")
    session.delete(task)
    session.commit()
    _audit("delete", "task", task_id, f"«{task.title}»", u, session)
    return RedirectResponse("/tasks", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Notifications
# ══════════════════════════════════════════════════════════════════

@app.get("/notifications", response_class=HTMLResponse)
def notifications_page(request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u:
        raise HTTPException(403)
    notifs = session.execute(
        select(Notification)
        .where(Notification.user_id == u.id)
        .order_by(desc(Notification.created_at)).limit(50)
    ).scalars().all()
    return templates.TemplateResponse(request, "notifications.html", {
        **_user_context(request), "notifications": notifs,
    })


@app.post("/notifications/read-all")
def notifications_read_all(request: Request, session: Session = Depends(get_db)):
    u = request.state.user
    if not u:
        raise HTTPException(403)
    session.execute(
        select(Notification)
        .where(Notification.user_id == u.id, Notification.is_read == False)
    )
    for n in session.execute(
        select(Notification)
        .where(Notification.user_id == u.id, Notification.is_read == False)
    ).scalars().all():
        n.is_read = True
    session.commit()
    return RedirectResponse("/notifications", status_code=303)


# ══════════════════════════════════════════════════════════════════
#  Schedule Calendar
# ══════════════════════════════════════════════════════════════════

@app.get("/schedule", response_class=HTMLResponse)
def schedule_calendar_page(request: Request, month: str = Query(""), session: Session = Depends(get_db)):
    today = datetime.now() + TIMEZONE_OFFSET
    if month:
        try:
            year, mon = map(int, month.split("-"))
            base = datetime(year, mon, 1)
        except (ValueError, TypeError):
            base = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        base = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    next_month = (base.replace(day=28) + timedelta(days=4)).replace(day=1)
    prev_month = (base - timedelta(days=1)).replace(day=1)

    orders = session.execute(
        select(Order).options(joinedload(Order.client), joinedload(Order.assignee))
        .where(
            or_(
                Order.scheduled_at.isnot(None),
                (Order.deadline.isnot(None)) & (Order.order_type == "print")
            ),
            or_(
                (Order.scheduled_at >= base) & (Order.scheduled_at < next_month),
                (Order.deadline >= base) & (Order.deadline < next_month) & (Order.order_type == "print")
            )
        )
        .order_by(Order.scheduled_at)
    ).unique().scalars().all()

    by_date = {}
    orders_json = {}
    for o in orders:
        if o.scheduled_at:
            d = o.scheduled_at.strftime("%Y-%m-%d")
            by_date.setdefault(d, []).append(o)
            orders_json.setdefault(d, []).append({
                "id": o.id, "client_name": o.client.full_name if o.client else "—",
                "printer": o.printer or "", "status": o.status,
                "order_type": o.order_type or "repair",
                "scheduled_at": o.scheduled_at.isoformat() if o.scheduled_at else "",
                "schedule_location": o.schedule_location or "",
            })
        if o.order_type == "print" and o.deadline:
            d = o.deadline.strftime("%Y-%m-%d")
            by_date.setdefault(d, []).append(o)
            if d not in orders_json or o.id not in [x["id"] for x in orders_json[d]]:
                orders_json.setdefault(d, []).append({
                    "id": o.id, "client_name": o.client.full_name if o.client else "—",
                    "printer": o.printer or "", "status": o.status,
                    "order_type": o.order_type or "repair",
                    "scheduled_at": o.deadline.isoformat() if o.deadline else "",
                    "schedule_location": o.schedule_location or "",
                })

    return templates.TemplateResponse(request, "schedule.html", {
        **_user_context(request),
        "orders_by_date": by_date, "orders_json": orders_json,
        "base_month": base, "next_month": next_month,
        "prev_month": prev_month, "today": today,
        "ORDER_STATUSES": ORDER_STATUSES, "timedelta": timedelta,
    })


# ══════════════════════════════════════════════════════════════════
#  API stubs (silence old CRM polling)
# ══════════════════════════════════════════════════════════════════

@app.get("/api/sse/events")
async def _sse_stub():
    return JSONResponse({"events": []})


@app.get("/api/dashboard")
async def _dashboard_stub():
    return JSONResponse({})


@app.get("/api/task-assignments/my")
async def _tasks_api(request: Request, session: Session = Depends(get_db)):
    u = _get_user_from_request(request)
    if not u:
        return JSONResponse([])
    tasks = session.execute(
        select(Task).options(joinedload(Task.creator), joinedload(Task.assignee))
        .where(Task.assigned_to == u.id, Task.status == "pending")
        .order_by(desc(Task.created_at))
    ).unique().scalars().all()
    return JSONResponse([{
        "id": t.id, "title": t.title, "description": t.description,
        "created_by": t.creator.full_name if t.creator else "—",
        "assigned_to": t.assignee.full_name if t.assignee else "—",
        "status": t.status, "created_at": t.created_at.isoformat(),
    } for t in tasks])


@app.get("/api/warehouse/products")
async def _wh_products_stub():
    return JSONResponse([])
